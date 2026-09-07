from __future__ import annotations

import json
import sys
import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from unittest import mock

import openpyxl
from fastapi import HTTPException


REPO_ROOT = Path(__file__).resolve().parents[2]
# Every test temp directory lives under one gitignored folder at the
# repository root, so a suite that dies before teardown cannot scatter
# tmp folders beside the code.
TEST_TEMP_ROOT = REPO_ROOT / "test"
TEST_TEMP_ROOT.mkdir(parents=True, exist_ok=True)
FRONTEND_ROOT = REPO_ROOT / "frontend"
PYTHON_API_SRC = REPO_ROOT / "python-api" / "src"
for path in (FRONTEND_ROOT, PYTHON_API_SRC):
    if str(path) not in sys.path:
        sys.path.insert(0, str(path))

from arcrho_api.dfm_contract import method_revisions, recalculate_dfm_method
from app_server.services import excel_link_service
from dependent_propagation_workspace_stub import IsolatedPropagationWorkspace


OLD_REFERENCE = "='C:\\Data\\[Book.xlsx]Sheet 1'!$A$1:$B$1"


def dfm_method_payload(inputs_first_cell: str) -> dict:
    return recalculate_dfm_method(
        {
            "details_tab": {
                "name": "Development",
                "output_type": "Selected Ultimate",
                "output_dataset": "Development Output",
                "input_triangle": "Paid",
                "origin_length": 12,
                "development_length": 12,
            },
            "ratios_tab": {
                "average_formulas": {
                    "label": ["User Entry"],
                    "custom_average_formula_settings": {"average_type": ["user_entry"]},
                    "selected": [[1, 1]],
                    "values": [[1.5, 1]],
                    "inputs": [[inputs_first_cell, "1"]],
                    "display_inputs": [[inputs_first_cell, "1"]],
                },
            },
            "results_tab": {},
        },
        input_snapshot={
            "name": "Paid",
            "data_format": "Triangle",
            "origin_labels": ["2024", "2025"],
            "development_labels": ["12", "24"],
            "values": [[100, 150], [200, None]],
            "mask": [[True, True], [True, False]],
            "number_format": "#,##0",
            "decimal_places": 0,
            "revision": "paid-r1",
        },
        timestamp="2026-01-01T00:00:00Z",
    )


class ExcelLinkReferenceRewriteTests(unittest.TestCase):
    def test_finds_quoted_inline_and_standalone_references(self) -> None:
        inline = excel_link_service.find_workbook_references(
            "1.5 * 'C:\\Data\\[Book.xlsx]Sheet 1'!$A$1 + 'D:\\Other\\[Second.xlsx]S2'!B2:C3"
        )
        self.assertEqual(
            [item["book_path"] for item in inline],
            ["C:\\Data\\Book.xlsx", "D:\\Other\\Second.xlsx"],
        )
        standalone = excel_link_service.find_workbook_references(
            "=C:\\Data\\[Book.xlsx]Sheet1!A1"
        )
        self.assertEqual(standalone[0]["book_path"], "C:\\Data\\Book.xlsx")
        self.assertEqual(excel_link_service.find_workbook_references("= \"Simple\" * 2"), [])

    def test_rewrites_only_matching_workbook_and_preserves_sheet_and_address(self) -> None:
        old_key = excel_link_service.workbook_key("c:\\data\\book.xlsx")
        text = "'C:\\Data\\[Book.xlsx]Sheet 1'!$A$1 + 'D:\\Other\\[Second.xlsx]S2'!B2"
        rewritten, changed = excel_link_service.rewrite_workbook_references(
            text, old_key, "E:\\Moved\\Book 2026.xlsx"
        )
        self.assertEqual(changed, 1)
        self.assertEqual(
            rewritten,
            "'E:\\Moved\\[Book 2026.xlsx]Sheet 1'!$A$1 + 'D:\\Other\\[Second.xlsx]S2'!B2",
        )

    def test_rewrites_standalone_unquoted_form_to_canonical_quoted_form(self) -> None:
        old_key = excel_link_service.workbook_key("C:\\Data\\Book.xlsx")
        rewritten, changed = excel_link_service.rewrite_workbook_references(
            "=C:\\Data\\[Book.xlsx]Sheet1!$A$1:$B$2", old_key, "E:\\Moved\\Book.xlsx"
        )
        self.assertEqual(changed, 1)
        self.assertEqual(rewritten, "='E:\\Moved\\[Book.xlsx]Sheet1'!$A$1:$B$2")

    def test_escapes_apostrophes_in_new_workbook_path(self) -> None:
        old_key = excel_link_service.workbook_key("C:\\Data\\Book.xlsx")
        rewritten, changed = excel_link_service.rewrite_workbook_references(
            "='C:\\Data\\[Book.xlsx]Sheet 1'!A1", old_key, "E:\\Ann's\\Book.xlsx"
        )
        self.assertEqual(changed, 1)
        self.assertEqual(rewritten, "='E:\\Ann''s\\[Book.xlsx]Sheet 1'!A1")
        round_trip = excel_link_service.find_workbook_references(rewritten)
        self.assertEqual(round_trip[0]["book_path"], "E:\\Ann's\\Book.xlsx")


class ExcelLinkFixture(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory(dir=TEST_TEMP_ROOT)
        root = Path(self.temp.name)
        self.sidecars = root / "sidecars"
        self.methods = root / "methods"
        self.datasets = root / "datasets"
        self.books = root / "books"
        for folder in (self.sidecars, self.methods, self.datasets, self.books):
            folder.mkdir()
        # The old workbook may be gone or unreadable; only its path matters.
        self.old_book = self.books / "Book.xlsx"
        self.old_book.write_bytes(b"old")
        # The new workbook is a real file: the retarget opens it where it runs.
        self.new_book = self.books / "Book 2026.xlsx"
        self.write_workbook(self.new_book, {("Sheet 1", "A1"): 100, ("Sheet 1", "B1"): 150})
        self.old_reference = (
            f"='{self.books}\\[Book.xlsx]Sheet 1'!$A$1:$B$1"
        )
        self.patchers = [
            IsolatedPropagationWorkspace(),
            mock.patch.object(
                excel_link_service.config,
                "get_project_dataset_sidecar_dir",
                return_value=str(self.sidecars),
            ),
            mock.patch.object(
                excel_link_service.config,
                "get_project_method_data_dir",
                return_value=str(self.methods),
            ),
        ]
        for patcher in self.patchers:
            patcher.start()

    def tearDown(self) -> None:
        for patcher in reversed(self.patchers):
            patcher.stop()
        self.temp.cleanup()

    def write_workbook(self, path: Path, cells: dict) -> None:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        for (sheet, cell), value in cells.items():
            if sheet not in workbook.sheetnames:
                workbook.create_sheet(sheet)
            workbook[sheet][cell] = value
        workbook.save(str(path))

    def write_json(self, path: Path, payload: dict) -> None:
        path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")

    def linked_sidecar(self, name: str = "Manual Paid") -> dict:
        return {
            "dataset_name": name,
            "dataset_type": name,
            "project_name": "Project",
            "reserving_class": "Class",
            "source_kind": "input",
            "data_format": "Triangle",
            "origin_length": 12,
            "development_length": 12,
            "stored_origin_length": 12,
            "stored_development_length": 12,
            "csv_file": f"{name}@12@12@cum@dev.csv",
            "updated_at": "2026-01-05T00:00:00Z",
            "audit_log": [
                {"event_date": "2026-01-05T00:00:00Z", "action": "Insert", "change_info": "", "user": "user1"},
            ],
            "external_links": [
                {
                    "reference": self.old_reference,
                    "target_cells": [
                        {"row": 0, "column": 0, "source_cell": "A1"},
                        {"row": 0, "column": 1, "source_cell": "B1"},
                    ],
                },
            ],
        }

    def write_dfm_method(self, first_input: str) -> dict:
        payload = dfm_method_payload(first_input)
        self.write_json(self.methods / "DFM@Development.json", payload)
        return payload


class ExcelLinkListingTests(ExcelLinkFixture):
    def test_list_groups_workbooks_across_datasets_and_dfm_methods(self) -> None:
        sidecar = self.linked_sidecar()
        # An instance whose name differs from its Dataset Type: the listing
        # carries both so the manager can open the exact instance.
        sidecar["dataset_type"] = "Paid Loss"
        self.write_json(self.sidecars / "Manual Paid.json", sidecar)
        self.write_json(self.sidecars / "No Links.json", {"dataset_name": "No Links"})
        (self.sidecars / "Broken.json").write_text("{not json", encoding="utf-8")
        self.write_dfm_method(f"'{self.books}\\[Book.xlsx]Sheet 1'!$A$1 * 2")
        self.write_json(self.methods / "DFM@Old.json", {"json_format": "arcrho-dfm-method-by-tab-v1"})

        listing = excel_link_service.list_reserving_class_excel_links("Project", "Class")

        self.assertTrue(listing["ok"])
        self.assertEqual(len(listing["workbooks"]), 1)
        workbook = listing["workbooks"][0]
        self.assertEqual(workbook["workbook_name"], "Book.xlsx")
        self.assertTrue(workbook["exists"])
        self.assertEqual(workbook["dataset_count"], 1)
        self.assertEqual(workbook["method_count"], 1)
        self.assertEqual(workbook["link_count"], 2)
        self.assertEqual(workbook["cell_count"], 3)
        self.assertEqual(
            [
                (item["kind"], item["name"], item["dataset_type"], item["method_type"])
                for item in workbook["usages"]
            ],
            # Method Type comes from the sidecar through its canonical
            # normalizer, so a manual input reads "None" exactly as the dataset
            # table shows it; a DFM usage is the method itself.
            [("dataset", "Manual Paid", "Paid Loss", "None"), ("dfm", "Development", "", "DFM")],
        )
        self.assertEqual(
            [item["file"] for item in listing["errors"]],
            ["Broken.json"],
        )

    def test_listing_answers_workbook_existence_from_this_host(self) -> None:
        # The listing is one hosted read: the host that scans the class is the
        # host that stats the workbooks, because it is the host that must open
        # them for any retarget. Existence never comes from a Client PC.
        self.write_json(self.sidecars / "Manual Paid.json", self.linked_sidecar())
        with mock.patch.object(
            excel_link_service.excel_service, "excel_workbook_properties_batch",
            wraps=excel_link_service.excel_service.excel_workbook_properties_batch,
        ) as stats:
            listing = excel_link_service.list_reserving_class_excel_links("Project", "Class")
        stats.assert_called_once_with([str(self.old_book)])
        self.assertTrue(listing["workbooks"][0]["exists"])
        self.assertIsNotNone(listing["workbooks"][0]["mtime"])

    def test_listing_carries_each_workbook_document_properties(self) -> None:
        # Created, Last Modified, and User are the workbook's own properties -
        # the workbook-side answer to the dataset table's three columns - read
        # on the host that holds the workbook.
        sidecar = self.linked_sidecar()
        sidecar["external_links"][0]["reference"] = (
            f"='{self.books}\\[Described.xlsx]Sheet 1'!$A$1"
        )
        self.write_json(self.sidecars / "Manual Paid.json", sidecar)
        described = self.books / "Described.xlsx"
        self.write_workbook(described, {("Sheet 1", "A1"): 1})
        workbook = openpyxl.load_workbook(str(described))
        workbook.properties.lastModifiedBy = "j.tanaka"
        workbook.properties.created = datetime(2024, 3, 2, 9, 15, 41)
        workbook.save(str(described))

        listing = excel_link_service.list_reserving_class_excel_links("Project", "Class")

        entry = listing["workbooks"][0]
        self.assertEqual(entry["workbook_name"], "Described.xlsx")
        self.assertEqual(entry["last_modified_by"], "j.tanaka")
        self.assertTrue(entry["created"].startswith("2024-03-02T09:15:41"))
        self.assertTrue(entry["modified"])

    def test_listing_leaves_properties_blank_for_a_non_package_workbook(self) -> None:
        # A legacy .xls or an encrypted package carries no readable properties.
        # It is still found and still relinkable, so the three columns are
        # blank rather than the workbook being reported missing.
        self.write_json(self.sidecars / "Manual Paid.json", self.linked_sidecar())

        listing = excel_link_service.list_reserving_class_excel_links("Project", "Class")

        entry = listing["workbooks"][0]
        self.assertTrue(entry["exists"])
        self.assertEqual(
            [entry["created"], entry["modified"], entry["last_modified_by"]],
            ["", "", ""],
        )

    def test_list_marks_missing_workbooks(self) -> None:
        sidecar = self.linked_sidecar()
        sidecar["external_links"][0]["reference"] = "='C:\\Gone\\[Missing.xlsx]S1'!A1:B1"
        self.write_json(self.sidecars / "Manual Paid.json", sidecar)

        listing = excel_link_service.list_reserving_class_excel_links("Project", "Class")

        self.assertEqual(listing["workbooks"][0]["workbook_name"], "Missing.xlsx")
        self.assertFalse(listing["workbooks"][0]["exists"])


class ExcelLinkRetargetTests(ExcelLinkFixture):
    """A retarget refreshes and re-saves every affected object from the new workbook."""

    def setUp(self) -> None:
        super().setUp()
        # Every canonical save enqueues its roots exactly like the real one, so
        # the deferred collector is exercised; the writes themselves are faked.
        def fake_dataset_save(project, reserving, name, **kwargs):
            # The real save persists the rewritten links; mimic that so the
            # inventory the response carries reflects the retarget.
            sidecar_path = self.sidecars / f"{name}.json"
            sidecar = json.loads(sidecar_path.read_text(encoding="utf-8"))
            sidecar["external_links"] = kwargs.get("external_links")
            self.write_json(sidecar_path, sidecar)
            propagation = excel_link_service.dependent_propagation_service.enqueue_save_propagation(
                project, reserving,
                [excel_link_service.dependent_propagation_service.changed_root(
                    name, kwargs.get("dataset_type") or name,
                )],
            )
            return {"ok": True, "calculated_updates": propagation, "propagation_ok": True}

        self.dataset_values = [[100, 150], [200, None]]
        # The shape the loader says it answered at, which is the dataset's own
        # display; None stands for a loader that names no shape at all.
        self.dataset_display_shape: tuple[int, int] | None = None
        self.retarget_patchers = [
            mock.patch.object(
                excel_link_service.dependent_propagation_service,
                "require_reserving_class_writable",
            ),
            mock.patch(
                "app_server.services.dataset_service.load_cached_dataset_values",
                side_effect=lambda *_a, **_k: {
                    "values": self.dataset_values,
                    **(
                        {
                            "origin_length": self.dataset_display_shape[0],
                            "development_length": self.dataset_display_shape[1],
                        }
                        if self.dataset_display_shape
                        else {}
                    ),
                },
            ),
            mock.patch(
                "app_server.services.dataset_service.save_dataset_sidecar",
                side_effect=fake_dataset_save,
            ),
            mock.patch(
                "app_server.services.dfm_service.save_dfm_method",
                return_value={"ok": True, "propagation": {"ok": True, "status": "unchanged"}, "propagation_ok": True},
            ),
            mock.patch("app_server.services.dfm_service._mark_review_needed"),
            mock.patch("app_server.services.dataset_instance_index_service.rebuild_index"),
        ]
        (
            self.engine, self.load_values, self.dataset_save, self.dfm_save,
            self.mark_review, self.rebuild,
        ) = [patcher.start() for patcher in self.retarget_patchers]

    def tearDown(self) -> None:
        for patcher in reversed(self.retarget_patchers):
            patcher.stop()
        super().tearDown()

    def run_retarget(self, new_book: Path | None = None) -> dict:
        submit = excel_link_service.dependent_propagation_service.submit_dependent_propagation_job
        with mock.patch.object(
            excel_link_service.dependent_propagation_service,
            "submit_dependent_propagation_job",
            side_effect=submit,
        ) as submitted:
            response = excel_link_service.retarget_reserving_class_workbook(
                "Project", "Class", str(self.old_book), str(new_book or self.new_book)
            )
        self.submitted = submitted
        return response

    def test_retarget_refuses_a_workbook_this_host_cannot_open(self) -> None:
        # The server's own view is the only one that counts; a path it cannot
        # open is refused before anything is written, with the reason.
        self.write_json(self.sidecars / "Manual Paid.json", self.linked_sidecar())
        before = (self.sidecars / "Manual Paid.json").read_bytes()
        for candidate in (self.books / "Nope.xlsx", self.old_book):
            with self.subTest(candidate=candidate.name):
                with self.assertRaises(HTTPException) as ctx:
                    excel_link_service.retarget_reserving_class_workbook(
                        "Project", "Class", str(self.books / "Other.xlsx"), str(candidate)
                    )
                self.assertEqual(ctx.exception.status_code, 400)
                self.assertTrue(
                    str(ctx.exception.detail).startswith(excel_link_service.WORKBOOK_UNREADABLE_PREFIX)
                )
        self.assertEqual((self.sidecars / "Manual Paid.json").read_bytes(), before)
        self.dataset_save.assert_not_called()

    def test_retarget_to_same_workbook_is_a_no_op(self) -> None:
        self.write_json(self.sidecars / "Manual Paid.json", self.linked_sidecar())
        before = (self.sidecars / "Manual Paid.json").read_bytes()
        response = excel_link_service.retarget_reserving_class_workbook(
            "Project", "Class", str(self.old_book), str(self.old_book).upper()
        )
        self.assertTrue(response["ok"])
        self.assertEqual(response["changed_file_count"], 0)
        self.assertIn("already the current link", response["message"])
        self.assertEqual((self.sidecars / "Manual Paid.json").read_bytes(), before)
        self.dataset_save.assert_not_called()

    def test_retarget_requires_a_live_engine_before_any_write(self) -> None:
        self.write_json(self.sidecars / "Manual Paid.json", self.linked_sidecar())
        before = (self.sidecars / "Manual Paid.json").read_bytes()
        self.engine.side_effect = HTTPException(503, "No ArcRho Engine instance is available.")
        with self.assertRaises(HTTPException) as ctx:
            self.run_retarget()
        self.assertEqual(ctx.exception.status_code, 503)
        self.assertEqual((self.sidecars / "Manual Paid.json").read_bytes(), before)
        self.dataset_save.assert_not_called()

    def test_retarget_saves_every_affected_dataset_even_when_values_match(self) -> None:
        # The new workbook holds exactly the stored numbers. The dataset is
        # still saved through the canonical flow: its link changed, that is an
        # audited change, and its dependents must be flagged for review.
        self.write_json(self.sidecars / "Manual Paid.json", self.linked_sidecar())
        untouched = {"dataset_name": "Other", "external_links": []}
        self.write_json(self.sidecars / "Other.json", untouched)
        untouched_bytes = (self.sidecars / "Other.json").read_bytes()

        response = self.run_retarget()

        self.assertTrue(response["ok"])
        self.assertEqual(response["changed_file_count"], 1)
        self.assertEqual(response["changed_link_count"], 1)
        self.assertEqual(response["refreshed_cell_count"], 2)
        self.assertEqual(response["failed_refresh_count"], 0)
        self.assertEqual(response["value_changed_file_count"], 0)
        self.engine.assert_called_once_with("Project", "Class")
        self.dataset_save.assert_called_once()
        kwargs = self.dataset_save.call_args.kwargs
        self.assertEqual(kwargs["values"], [[100, 150], [200, None]])
        self.assertIn("Book 2026.xlsx", kwargs["external_links"][0]["reference"])
        self.assertEqual(kwargs["origin_length"], 12)
        self.assertEqual((self.sidecars / "Other.json").read_bytes(), untouched_bytes)
        # The response carries the refreshed inventory, resolved on this host.
        self.assertEqual(response["workbooks"][0]["workbook_name"], "Book 2026.xlsx")
        self.assertTrue(response["workbooks"][0]["exists"])

    def test_retarget_writes_the_csv_back_at_the_shape_it_read(self) -> None:
        # A sidecar naming a display no roll-up can build is answered at the
        # file's own shape, and the refreshed values go back there.
        sidecar = self.linked_sidecar()
        sidecar["origin_length"] = 36
        sidecar["development_length"] = 36
        self.write_json(self.sidecars / "Manual Paid.json", sidecar)

        self.assertTrue(self.run_retarget()["ok"])

        kwargs = self.dataset_save.call_args.kwargs
        self.assertEqual((kwargs["origin_length"], kwargs["development_length"]), (12, 12))

    def test_retarget_reads_a_finer_store_at_the_display_its_links_name(self) -> None:
        # A link names a cell of the grid its dataset was displayed at when it
        # was written, which is not the file's own grid when the file is stored
        # finer. The values are read at that display and handed back the same
        # way, and the save puts them into the store as any save from that
        # view does.
        sidecar = self.linked_sidecar()
        sidecar["stored_development_length"] = 1
        sidecar["csv_file"] = "Manual Paid@12@1@cum@dev.csv"
        self.write_json(self.sidecars / "Manual Paid.json", sidecar)
        self.dataset_display_shape = (12, 12)

        self.assertTrue(self.run_retarget()["ok"])

        self.assertIs(self.load_values.call_args.kwargs["at_linked_shape"], True)
        kwargs = self.dataset_save.call_args.kwargs
        self.assertEqual((kwargs["origin_length"], kwargs["development_length"]), (12, 12))
        self.assertEqual(kwargs["display_at"], (12, 12))

    def test_retarget_leaves_a_display_that_moved_on_from_the_links(self) -> None:
        # The dataset has been saved at a yearly-of-three view since its links
        # were written at 12/12. The links are still read and written at 12/12,
        # and the save records the 36/36 display rather than pulling it back.
        sidecar = self.linked_sidecar()
        sidecar["origin_length"] = 36
        sidecar["development_length"] = 36
        sidecar["linked_origin_length"] = 12
        sidecar["linked_development_length"] = 12
        self.write_json(self.sidecars / "Manual Paid.json", sidecar)
        self.dataset_display_shape = (12, 12)

        self.assertTrue(self.run_retarget()["ok"])

        self.assertIs(self.load_values.call_args.kwargs["at_linked_shape"], True)
        kwargs = self.dataset_save.call_args.kwargs
        self.assertEqual((kwargs["origin_length"], kwargs["development_length"]), (12, 12))
        self.assertEqual(kwargs["display_at"], (36, 36))

    def test_retarget_reads_the_new_workbook_and_applies_changed_values(self) -> None:
        self.write_json(self.sidecars / "Manual Paid.json", self.linked_sidecar())
        self.write_workbook(self.new_book, {("Sheet 1", "A1"): 111.5, ("Sheet 1", "B1"): None})

        response = self.run_retarget()

        self.assertEqual(response["value_changed_file_count"], 1)
        kwargs = self.dataset_save.call_args.kwargs
        self.assertEqual(kwargs["values"][0][0], 111.5)
        self.assertIsNone(kwargs["values"][0][1])
        self.assertEqual(kwargs["values"][1], [200, None])

    def test_retarget_save_preserves_zero_decimal_places(self) -> None:
        sidecar = self.linked_sidecar()
        sidecar["decimal_places"] = 0
        self.write_json(self.sidecars / "Manual Paid.json", sidecar)
        self.run_retarget()
        self.assertEqual(self.dataset_save.call_args.kwargs["decimal_places"], 0)

    def test_refresh_is_per_link_atomic_and_still_retargets_failed_links(self) -> None:
        sidecar = self.linked_sidecar()
        sidecar["external_links"].append({
            "reference": f"='{self.books}\\[Book.xlsx]Sheet 1'!$C$3",
            "target_cells": [{"row": 1, "column": 0, "source_cell": "C3"}],
        })
        self.write_json(self.sidecars / "Manual Paid.json", sidecar)
        self.write_workbook(self.new_book, {
            ("Sheet 1", "A1"): 111.5, ("Sheet 1", "B1"): 152, ("Sheet 1", "C3"): "text",
        })

        response = self.run_retarget()

        result = response["results"][0]
        self.assertTrue(result["ok"])
        self.assertEqual(result["changed_link_count"], 2)
        self.assertEqual(result["refreshed_cell_count"], 2)
        self.assertEqual(result["failed_refresh_count"], 1)
        self.assertIn("C3", result["refresh_errors"][0])
        kwargs = self.dataset_save.call_args.kwargs
        self.assertEqual(kwargs["values"][0], [111.5, 152])
        self.assertEqual(kwargs["values"][1], [200, None])
        self.assertIn("Book 2026.xlsx", kwargs["external_links"][1]["reference"])

    def test_retarget_saves_the_sidecar_alone_when_the_csv_cannot_be_loaded(self) -> None:
        self.write_json(self.sidecars / "Manual Paid.json", self.linked_sidecar())
        self.load_values.side_effect = HTTPException(404, "Dataset CSV not found.")

        response = self.run_retarget()

        result = response["results"][0]
        self.assertTrue(result["ok"])
        self.assertEqual(result["failed_refresh_count"], 2)
        self.assertIn("Dataset CSV not found.", result["refresh_errors"][0])
        kwargs = self.dataset_save.call_args.kwargs
        self.assertIsNone(kwargs["values"])
        self.assertIn("Book 2026.xlsx", kwargs["external_links"][0]["reference"])
        self.assertEqual(kwargs["origin_length"], 12)

    def test_retarget_saves_and_flags_a_dfm_and_walks_from_its_output(self) -> None:
        method = self.write_dfm_method(f"'{self.books}\\[Book.xlsx]Sheet 1'!$A$1 * 2")
        self.write_workbook(self.new_book, {("Sheet 1", "A1"): 2})

        response = self.run_retarget()

        result = response["results"][0]
        self.assertEqual(result["kind"], "dfm")
        self.assertTrue(result["value_changed"])
        self.assertEqual(result["refreshed_cell_count"], 1)
        self.assertEqual(result["output_dataset"], "Development Output")
        self.dfm_save.assert_called_once()
        merged = self.dfm_save.call_args.args[2]
        formulas = merged["ratios_tab"]["average_formulas"]
        self.assertEqual(formulas["values"][0][0], 4)
        self.assertIn("Book 2026.xlsx", formulas["inputs"][0][0])
        self.assertEqual(
            self.dfm_save.call_args.kwargs["expected_owned_revision"],
            method_revisions(method)["owned_revision"],
        )
        # An explicit save resets the output to current; the retarget then
        # flags it, and the output is a walk root even though the mocked save
        # reported an unchanged publication.
        self.mark_review.assert_called_once_with("Project", "Class", "Development Output")
        self.submitted.assert_called_once()
        self.assertEqual(
            self.submitted.call_args.args[2],
            [{"dataset_name": "Development Output", "dataset_type": "Selected Ultimate"}],
        )
        self.assertEqual(response["propagation"]["status"], "queued")
        self.assertTrue(response["propagation_ok"])

    def test_refresh_rejects_nonpositive_dfm_results_and_keeps_stored_values(self) -> None:
        self.write_dfm_method(f"'{self.books}\\[Book.xlsx]Sheet 1'!$A$1 * 2")
        self.write_workbook(self.new_book, {("Sheet 1", "A1"): -3})

        response = self.run_retarget()

        result = response["results"][0]
        self.assertFalse(result["value_changed"])
        self.assertEqual(result["failed_refresh_count"], 1)
        self.assertIn("greater than 0", result["refresh_errors"][0])
        merged = self.dfm_save.call_args.args[2]
        self.assertEqual(merged["ratios_tab"]["average_formulas"]["values"][0][0], 1.5)
        self.assertIn("Book 2026.xlsx", merged["ratios_tab"]["average_formulas"]["inputs"][0][0])
        self.mark_review.assert_called_once()

    def test_refresh_spills_dfm_ranges_into_literal_non_anchor_cells(self) -> None:
        self.write_dfm_method(f"='{self.books}\\[Book.xlsx]Sheet 1'!$A$1:$B$1")
        self.write_workbook(self.new_book, {("Sheet 1", "A1"): 1.1, ("Sheet 1", "B1"): 1.2})

        response = self.run_retarget()

        result = response["results"][0]
        self.assertTrue(result["value_changed"])
        self.assertEqual(result["refreshed_cell_count"], 2)
        merged = self.dfm_save.call_args.args[2]
        formulas = merged["ratios_tab"]["average_formulas"]
        self.assertEqual(formulas["values"][0], [1.1, 1.2])
        self.assertIn("Book 2026.xlsx", formulas["inputs"][0][0])
        self.assertEqual(formulas["inputs"][0][1], "1.2")
        self.assertEqual(formulas["display_inputs"][0][1], "1.2")

    def test_nested_saves_share_one_walk_and_the_index_is_rebuilt_once_after_it(self) -> None:
        # Two datasets and one DFM: three canonical saves, one propagation
        # submission carrying every root, then one index rebuild.
        self.write_json(self.sidecars / "Manual Paid.json", self.linked_sidecar())
        self.write_json(self.sidecars / "Manual Incurred.json", self.linked_sidecar("Manual Incurred"))
        self.write_dfm_method(f"'{self.books}\\[Book.xlsx]Sheet 1'!$A$1 * 2")

        response = self.run_retarget()

        self.assertEqual(response["changed_file_count"], 3)
        self.assertEqual(self.dataset_save.call_count, 2)
        self.submitted.assert_called_once()
        self.assertEqual(
            [root["dataset_name"] for root in self.submitted.call_args.args[2]],
            ["Manual Incurred", "Manual Paid", "Development Output"],
        )
        self.rebuild.assert_called_once_with("Project", "Class")
        self.assertEqual(
            sorted((item["kind"], item["name"]) for item in response["results"]),
            [("dataset", "Manual Incurred"), ("dataset", "Manual Paid"), ("dfm", "Development")],
        )

    def test_a_dfm_that_fails_to_save_is_reported_and_the_rest_still_land(self) -> None:
        self.write_json(self.sidecars / "Manual Paid.json", self.linked_sidecar())
        self.write_dfm_method(f"'{self.books}\\[Book.xlsx]Sheet 1'!$A$1 * 2")
        self.dfm_save.side_effect = HTTPException(409, "DFM changed on disk.")

        response = self.run_retarget()

        self.assertFalse(response["ok"])
        self.assertEqual(response["changed_file_count"], 1)
        failed = [item for item in response["results"] if not item["ok"]]
        self.assertEqual(failed[0]["name"], "Development")
        self.assertEqual(failed[0]["error"], "DFM changed on disk.")
        self.mark_review.assert_not_called()

    def test_save_propagation_roots_mirror_the_retarget(self) -> None:
        self.write_json(self.sidecars / "Manual Paid.json", self.linked_sidecar())
        self.write_json(self.sidecars / "Other.json", {"dataset_name": "Other", "external_links": []})
        self.write_dfm_method(f"'{self.books}\\[Book.xlsx]Sheet 1'!$A$1 * 2")

        roots = excel_link_service.save_propagation_roots(
            "Project", "Class", str(self.old_book), str(self.new_book)
        )
        self.assertEqual(
            roots,
            [("Manual Paid", "Manual Paid"), ("Development Output", "Selected Ultimate")],
        )
        self.assertEqual(
            excel_link_service.save_propagation_roots(
                "Project", "Class", str(self.old_book), str(self.old_book)
            ),
            [],
        )


class ExcelLinkHostedSaveRegistrationTests(unittest.TestCase):
    def test_the_retarget_is_a_hosted_save_kind_owned_by_this_service(self) -> None:
        from arcrho_engine_save_contract import SAVE_JOB_KINDS

        self.assertEqual(
            SAVE_JOB_KINDS["excel_link_retarget"],
            ("excel_link_service", "retarget_reserving_class_workbook"),
        )


if __name__ == "__main__":
    unittest.main()
