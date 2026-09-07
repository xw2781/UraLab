from __future__ import annotations

import sys
import stat
import tempfile
import unittest
from concurrent.futures import Future
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace
from unittest import mock


REPO_ROOT = Path(__file__).resolve().parents[2]
# Every test temp directory lives under one gitignored folder at the
# repository root, so a suite that dies before teardown cannot scatter
# tmp folders beside the code.
TEST_TEMP_ROOT = REPO_ROOT / "test"
TEST_TEMP_ROOT.mkdir(parents=True, exist_ok=True)
FRONTEND_ROOT = REPO_ROOT / "frontend"
if str(FRONTEND_ROOT) not in sys.path:
    sys.path.insert(0, str(FRONTEND_ROOT))

import openpyxl
from openpyxl.worksheet._read_only import ReadOnlyWorksheet

from app_server.services import excel_service


class _Workbook:
    """A workbook whose every cell holds one value, read the way the real one is.

    The service reads a sheet by walking the rectangle its requested cells
    span, so the fake answers ``iter_rows`` rather than single addresses.
    """

    sheetnames = ["Sheet1"]

    def __init__(self, value: float) -> None:
        self.value = value
        self.closed = False

    def __getitem__(self, key: str):
        if key == "Sheet1":
            return self
        raise KeyError(key)

    def iter_rows(self, min_row, max_row, min_col, max_col, values_only=False):
        del values_only
        for _row in range(min_row, max_row + 1):
            yield tuple(self.value for _column in range(min_col, max_col + 1))

    def close(self) -> None:
        self.closed = True


class _RecordingExecutor:
    max_workers = 0

    def __init__(self, *, max_workers: int, thread_name_prefix: str) -> None:
        del thread_name_prefix
        type(self).max_workers = max_workers

    def __enter__(self):
        return self

    def __exit__(self, *_args) -> None:
        return None

    def submit(self, fn, *args):
        future = Future()
        future.set_result(fn(*args))
        return future


class ExcelBatchReadTests(unittest.TestCase):
    def work_dir(self) -> Path:
        """A scratch folder for the tests that need a workbook on disk.

        Created per test rather than in ``setUp`` so the tests that mock the
        file system away keep touching no files at all, and inside the
        repository so a validation run never writes outside it.
        """

        temp_dir = tempfile.TemporaryDirectory(dir=TEST_TEMP_ROOT)
        self.addCleanup(temp_dir.cleanup)
        return Path(temp_dir.name)

    def item(self, book: str, cell: str = "A1"):
        return SimpleNamespace(book_path=book, sheet="Sheet1", cell=cell)

    def test_deduplicates_cells_opens_each_workbook_once_and_preserves_order(self) -> None:
        first = _Workbook(12.5)
        second = _Workbook(33.0)

        # Keyed on the path, not an ordered side_effect list: the two workbooks
        # are opened on separate threads, so their order is not the caller's.
        def fake_load(path, **_kwargs):
            return first if "first" in str(path) else second

        items = [
            self.item("first.xlsx"),
            self.item("second.xlsx"),
            self.item("first.xlsx"),
        ]
        with (
            mock.patch.object(Path, "exists", return_value=True),
            mock.patch.object(
                excel_service.openpyxl, "load_workbook", side_effect=fake_load
            ) as load,
        ):
            result = excel_service.excel_read_cells_batch(items)

        self.assertTrue(result["ok"])
        self.assertEqual(result["results"], [
            {"ok": True, "value": 12.5},
            {"ok": True, "value": 33.0},
            {"ok": True, "value": 12.5},
        ])
        self.assertEqual(load.call_count, 2)
        self.assertTrue(first.closed)
        self.assertTrue(second.closed)

    def test_a_cell_that_looks_empty_is_a_blank_value_not_an_error(self) -> None:
        # Untouched, outside the used range, an empty string cached by a
        # formula, and a whitespace-only cell all read as the blank ArcRho
        # stores as null. Only text that is not a number - the #REF! a deleted
        # row leaves behind - is that cell's own error.
        for raw in (None, "", " ", "\u00a0", "\t"):
            self.assertEqual(
                excel_service.workbook_cell_value(raw), {"ok": True, "value": None}
            )
        self.assertEqual(excel_service.workbook_cell_value(0), {"ok": True, "value": 0.0})
        self.assertEqual(excel_service.workbook_cell_value(-2.5), {"ok": True, "value": -2.5})
        broken = excel_service.workbook_cell_value("#REF!")
        self.assertFalse(broken["ok"])
        self.assertIn("#REF!", broken["error"])

    def test_blank_cells_inside_a_range_do_not_fail_the_read(self) -> None:
        # A linked range is applied whole or not at all, so one blank inside it
        # rejecting the read would drop every value in the range.
        work = self.work_dir()
        book = work / "Range.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet["A1"] = 1
        sheet["A3"] = 3
        sheet["B4"] = 4
        workbook.save(book)

        items = [
            SimpleNamespace(book_path=str(book), sheet="Sheet1", cell=cell)
            for cell in ("A1", "A2", "A3", "A4", "Z99")
        ]
        result = excel_service.excel_read_cells_batch(items)

        self.assertTrue(all(item["ok"] for item in result["results"]))
        self.assertEqual(
            [item["value"] for item in result["results"]], [1.0, None, 3.0, None, None]
        )

    def test_a_linked_range_walks_its_sheet_once(self) -> None:
        # A read-only worksheet re-reads the sheet from its first row every
        # time it is asked for a cell by address, so reading a linked range one
        # address at a time costs one pass per cell and takes minutes on a
        # 120x120 range. The whole rectangle is answered by a single walk.
        work = self.work_dir()
        book = work / "Grid.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        for row in range(1, 31):
            for column in range(1, 31):
                sheet.cell(row=row, column=column, value=row * 100 + column)
        workbook.save(book)

        items = [
            SimpleNamespace(book_path=str(book), sheet="Sheet1", cell=f"{letter}{row}")
            for row in range(1, 31)
            for letter in (chr(ord("A") + index) for index in range(30 - 4))
        ]
        walks = []
        original = ReadOnlyWorksheet.iter_rows

        def counting_iter_rows(worksheet, *args, **kwargs):
            walks.append(kwargs)
            return original(worksheet, *args, **kwargs)

        with mock.patch.object(ReadOnlyWorksheet, "iter_rows", counting_iter_rows):
            result = excel_service.excel_read_cells_batch(items)

        self.assertEqual(len(walks), 1)
        self.assertTrue(all(item["ok"] for item in result["results"]))
        self.assertEqual(result["results"][0]["value"], 101.0)
        self.assertEqual(result["results"][-1]["value"], 3026.0)

    def test_each_sheet_of_a_workbook_is_walked_once(self) -> None:
        # Cells of two sheets are two walks of one open workbook, not one walk
        # per cell and not one open per sheet.
        work = self.work_dir()
        book = work / "Sheets.xlsx"
        workbook = openpyxl.Workbook()
        first = workbook.active
        first.title = "Sheet1"
        first["A1"] = 1
        first["A2"] = 2
        second = workbook.create_sheet("Sheet2")
        second["B1"] = 3
        workbook.save(book)

        items = [
            SimpleNamespace(book_path=str(book), sheet=sheet, cell=cell)
            for sheet, cell in (("Sheet1", "A1"), ("Sheet2", "B1"), ("Sheet1", "A2"))
        ]
        walks = []
        original = ReadOnlyWorksheet.iter_rows

        def counting_iter_rows(worksheet, *args, **kwargs):
            walks.append(worksheet.title)
            return original(worksheet, *args, **kwargs)

        with (
            mock.patch.object(ReadOnlyWorksheet, "iter_rows", counting_iter_rows),
            mock.patch.object(
                excel_service.openpyxl, "load_workbook", wraps=openpyxl.load_workbook
            ) as load,
        ):
            result = excel_service.excel_read_cells_batch(items)

        self.assertEqual(sorted(walks), ["Sheet1", "Sheet2"])
        self.assertEqual(load.call_count, 1)
        self.assertEqual([item["value"] for item in result["results"]], [1.0, 3.0, 2.0])

    def test_a_read_leaves_the_workbook_file_unlocked(self) -> None:
        # The workbook belongs to whoever is editing it in Excel: a link read
        # that held the file open would stop them saving it.
        work = self.work_dir()
        book = work / "Free.xlsx"
        workbook = openpyxl.Workbook()
        workbook.active.title = "Sheet1"
        workbook.active["A1"] = 5
        workbook.save(book)

        excel_service.excel_read_cell(str(book), "Sheet1", "A1")
        excel_service.excel_read_cells_batch(
            [SimpleNamespace(book_path=str(book), sheet="Sheet1", cell="A1")]
        )

        book.unlink()
        self.assertFalse(book.exists())

    def test_single_and_batch_reads_answer_one_cell_identically(self) -> None:
        # One rule for what a linked cell means, so a client commit and the
        # hosted retarget cannot disagree about the same workbook cell.
        work = self.work_dir()
        book = work / "One.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet["A2"] = "#REF!"
        sheet["A3"] = 7
        workbook.save(book)

        for cell in ("A1", "A2", "A3"):
            single = excel_service.excel_read_cell(str(book), "Sheet1", cell)
            batch = excel_service.excel_read_cells_batch(
                [SimpleNamespace(book_path=str(book), sheet="Sheet1", cell=cell)]
            )["results"][0]
            self.assertEqual(single, batch, cell)

    def test_one_bad_cell_does_not_poison_the_rest_of_its_workbook(self) -> None:
        # A saved link is validated cell by cell, so an address the sheet cannot
        # resolve names itself instead of failing every cell read from the same
        # workbook - which is the difference between "this reference broke" and
        # "40 cells failed".
        class _MixedWorkbook:
            sheetnames = ["Sheet1"]

            def __init__(self) -> None:
                self.closed = False

            def __getitem__(self, key: str):
                if key == "Sheet1":
                    return self
                raise KeyError(key)

            def iter_rows(self, min_row, max_row, min_col, max_col, values_only=False):
                del values_only
                values = {1: 7.5, 3: "#REF!"}
                for _row in range(min_row, max_row + 1):
                    yield tuple(
                        values.get(column) for column in range(min_col, max_col + 1)
                    )

            def close(self) -> None:
                self.closed = True

        book = _MixedWorkbook()
        # "B1:B2" is a range, not a cell address: a sidecar written by hand can
        # hold one, and it is the whole of what "unreadable address" now means.
        items = [self.item("first.xlsx", cell) for cell in ("A1", "B1:B2", "C1")]
        with (
            mock.patch.object(Path, "exists", return_value=True),
            mock.patch.object(
                excel_service.openpyxl, "load_workbook", return_value=book
            ) as load,
        ):
            result = excel_service.excel_read_cells_batch(items)

        good, unreadable, not_numeric = result["results"]
        self.assertEqual(good, {"ok": True, "value": 7.5})
        self.assertFalse(unreadable["ok"])
        self.assertIn("B1:B2", unreadable["error"])
        self.assertFalse(not_numeric["ok"])
        self.assertIn("#REF!", not_numeric["error"])
        self.assertEqual(load.call_count, 1)
        self.assertTrue(book.closed)

    def test_validate_links_answers_cells_and_workbook_times_in_one_pass(self) -> None:
        # The open dataset asks two questions - is every reference still
        # readable, and is any workbook newer - and pays one pass for both.
        first = _Workbook(12.5)
        second = _Workbook(33.0)
        stat_calls = []

        def fake_stat(path: str):
            stat_calls.append(path)
            return SimpleNamespace(
                st_mtime=101.0 if "first" in path else 202.0,
                st_mode=stat.S_IFREG,
            )

        # Keyed on the path, not an ordered side_effect list: the two workbooks
        # are opened on separate threads, so their order is not the caller's.
        def fake_load(path, **_kwargs):
            return first if "first" in str(path) else second

        items = [
            self.item("first.xlsx"),
            self.item("second.xlsx"),
            self.item("first.xlsx", "B2"),
        ]
        with (
            mock.patch.object(Path, "exists", return_value=True),
            mock.patch.object(
                excel_service.openpyxl, "load_workbook", side_effect=fake_load
            ) as load,
            mock.patch.object(excel_service.os, "stat", side_effect=fake_stat),
        ):
            result = excel_service.excel_validate_links(items)

        self.assertTrue(result["ok"])
        self.assertEqual(
            [item["value"] for item in result["results"]], [12.5, 33.0, 12.5]
        )
        # One open and one stat per distinct workbook, however many cells asked.
        self.assertEqual(load.call_count, 2)
        self.assertEqual(len(stat_calls), 2)
        self.assertEqual(
            sorted(workbook["mtime"] for workbook in result["workbooks"]),
            [101.0, 202.0],
        )
        self.assertTrue(all(workbook["ok"] for workbook in result["workbooks"]))

    def test_validate_links_reports_a_workbook_it_cannot_reach(self) -> None:
        # A workbook that moved fails both halves of the answer, and the caller
        # needs both: the reference is broken, and its timestamp is unknown.
        missing = self.work_dir() / "Gone.xlsx"

        result = excel_service.excel_validate_links(
            [SimpleNamespace(book_path=str(missing), sheet="Sheet1", cell="A1")]
        )

        self.assertFalse(result["results"][0]["ok"])
        self.assertIn("File not found", result["results"][0]["error"])
        self.assertEqual(len(result["workbooks"]), 1)
        self.assertFalse(result["workbooks"][0]["ok"])

    def test_read_cells_batch_does_not_pay_for_workbook_timestamps(self) -> None:
        # Only the validating read stats the workbooks; the plain batch read
        # every commit and refresh uses must not start doing it too.
        with (
            mock.patch.object(Path, "exists", return_value=True),
            mock.patch.object(
                excel_service.openpyxl, "load_workbook", return_value=_Workbook(1.0)
            ),
            mock.patch.object(excel_service.os, "stat") as stat_call,
        ):
            result = excel_service.excel_read_cells_batch([self.item("first.xlsx")])

        self.assertNotIn("workbooks", result)
        stat_call.assert_not_called()

    def test_bounds_workbook_concurrency(self) -> None:
        items = [self.item(f"missing-{index}.xlsx") for index in range(8)]
        with mock.patch.object(excel_service, "ThreadPoolExecutor", _RecordingExecutor):
            result = excel_service.excel_read_cells_batch(items)

        self.assertEqual(_RecordingExecutor.max_workers, excel_service.EXCEL_BATCH_MAX_WORKERS)
        self.assertEqual(len(result["results"]), len(items))
        self.assertTrue(all(not item["ok"] for item in result["results"]))

    def test_file_mtimes_deduplicate_stat_calls_and_preserve_order(self) -> None:
        calls = []

        def fake_stat(path: str):
            calls.append(path)
            return SimpleNamespace(
                st_mtime=101.0 if "first" in path else 202.0,
                st_mode=stat.S_IFREG,
            )

        paths = ["first.xlsx", "second.xlsx", "first.xlsx", ""]
        with (
            mock.patch.object(excel_service.os, "stat", side_effect=fake_stat),
            mock.patch.object(excel_service, "ThreadPoolExecutor", _RecordingExecutor),
        ):
            result = excel_service.excel_file_mtimes_batch(paths)

        self.assertTrue(result["ok"])
        self.assertEqual([item.get("mtime") for item in result["results"][:3]], [101.0, 202.0, 101.0])
        self.assertFalse(result["results"][3]["ok"])
        self.assertEqual(len(calls), 2)
        self.assertEqual(_RecordingExecutor.max_workers, 3)

    def test_file_mtime_concurrency_is_bounded(self) -> None:
        paths = [f"workbook-{index}.xlsx" for index in range(8)]
        with (
            mock.patch.object(
                excel_service.os,
                "stat",
                return_value=SimpleNamespace(st_mtime=100.0, st_mode=stat.S_IFREG),
            ),
            mock.patch.object(excel_service, "ThreadPoolExecutor", _RecordingExecutor),
        ):
            result = excel_service.excel_file_mtimes_batch(paths)

        self.assertEqual(_RecordingExecutor.max_workers, excel_service.EXCEL_BATCH_MAX_WORKERS)
        self.assertEqual(len(result["results"]), len(paths))

    def test_workbook_properties_read_the_package_core_properties(self) -> None:
        # Created, Last Modified, and User are the workbook's own record of
        # itself - the workbook-side answer to the dataset table's columns -
        # so they come from docProps/core.xml, not from the file system.
        work = self.work_dir()
        book = work / "Book.xlsx"
        workbook = openpyxl.Workbook()
        workbook.properties.creator = "a.author"
        workbook.properties.lastModifiedBy = "j.tanaka"
        workbook.properties.created = datetime(2024, 3, 2, 9, 15, 41)
        workbook.save(book)

        # Not an OOXML package: a legacy .xls, an encrypted workbook, anything
        # that is not a readable zip. It is still a good link target, so it is
        # reported without properties rather than as an error.
        legacy = work / "Legacy.xls"
        legacy.write_bytes(bytes([0xd0, 0xcf, 0x11, 0xe0]) + b" not a zip")
        missing = work / "Gone.xlsx"

        result = excel_service.excel_workbook_properties_batch(
            [str(book), str(legacy), str(missing), "", str(book)]
        )

        good, plain, absent, blank, repeat = result["results"]
        self.assertTrue(good["ok"])
        self.assertEqual(good["last_modified_by"], "j.tanaka")
        self.assertTrue(good["created"].startswith("2024-03-02T09:15:41"))
        self.assertTrue(good["modified"])
        self.assertIsNotNone(good["mtime"])
        self.assertTrue(plain["ok"])
        self.assertNotIn("last_modified_by", plain)
        self.assertFalse(absent["ok"])
        self.assertFalse(blank["ok"])
        # One read per distinct path; the repeat carries the same answer.
        self.assertEqual(repeat, good)

    def test_file_mtimes_batch_keeps_its_own_result_shape(self) -> None:
        # The properties batch is the richer read; the mtimes batch every
        # freshness check uses must not start paying for a docProps read.
        work = self.work_dir()
        book = work / "Book.xlsx"
        openpyxl.Workbook().save(book)

        result = excel_service.excel_file_mtimes_batch([str(book)])

        self.assertEqual(sorted(result["results"][0].keys()), ["mtime", "ok", "path"])


if __name__ == "__main__":
    unittest.main()
