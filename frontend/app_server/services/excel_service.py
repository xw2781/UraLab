"""Workbook reads (openpyxl) and the one COM interop action (open in Excel).

Cell values and file stats are plain file reads of the workbook — no Excel
installation is involved — so they run wherever the workbook is reachable.
Only ``excel_open_workbook`` drives a desktop Excel through win32com.
"""
from __future__ import annotations

import os
import stat
import zipfile
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from typing import Any, Dict, List, Mapping, Tuple
from xml.etree import ElementTree

import openpyxl
from openpyxl.utils.cell import coordinate_to_tuple


EXCEL_BATCH_MAX_WORKERS = 4
# Identifies one requested cell: normalized workbook path, sheet, and address.
CellKey = Tuple[str, str, str]
# The OOXML package part that carries a workbook's own Created, Modified, and
# Last saved by properties, and the namespaces its elements are written in.
_CORE_PROPERTIES_ENTRY = "docProps/core.xml"
_CORE_PROPERTIES_NS = {
    "cp": "http://schemas.openxmlformats.org/package/2006/metadata/core-properties",
    "dcterms": "http://purl.org/dc/terms/",
}
# The part holds a handful of short strings; a larger one is a malformed or
# hostile package rather than something worth parsing.
_CORE_PROPERTIES_MAX_BYTES = 64 * 1024


def _item_field(item: Any, name: str) -> Any:
    """Read one field of a batch item, which is a router model or a plain dict."""

    if isinstance(item, Mapping):
        return item.get(name)
    return getattr(item, name, None)


def excel_workbook_readable(book_path: str) -> Dict[str, Any]:
    """Report whether this process can open the workbook as an Excel file.

    Answers for the machine it runs on: on ArcRho Server this is the truth
    the Excel Link Manager reports and the retarget requires, and a path the
    server cannot reach is refused rather than guessed at.
    """

    book = Path(str(book_path or "").strip())
    if not str(book):
        return {"ok": False, "error": "Workbook path is empty."}
    try:
        if not book.is_file():
            return {"ok": False, "error": "The workbook was not found."}
        wb = openpyxl.load_workbook(str(book), data_only=True, read_only=True)
        wb.close()
    except PermissionError:
        return {"ok": False, "error": "The workbook is not readable (permission denied)."}
    except OSError as exc:
        return {"ok": False, "error": f"The workbook could not be opened: {exc.strerror or exc}"}
    except Exception as exc:  # openpyxl raises its own zip/format errors
        return {"ok": False, "error": f"The workbook could not be opened as an Excel file: {exc}"}
    return {"ok": True}


def workbook_cell_value(raw: Any) -> Dict[str, Any]:
    """Turn one raw workbook cell into the value ArcRho stores, or its error.

    This is the single rule for what a linked cell means, shared by the single
    read and the batch reads so a client and the hosted retarget can never
    disagree about one workbook cell.

    A cell the user sees as empty is a blank value (``None``), whether it was
    never touched, sits outside the sheet's used range, or holds a formula whose
    cached result is an empty or whitespace-only string - Excel writes that last
    case as text, and rejecting it would fail a whole linked range over cells
    that look empty on screen. Text that is not a number - the ``#REF!`` a
    deleted row leaves behind - is still that cell's own error.
    """

    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return {"ok": True, "value": None}
    try:
        return {"ok": True, "value": float(raw)}
    except (TypeError, ValueError):
        return {"ok": False, "error": f"Not numeric: {raw!r}"}


def excel_read_cell(book_path: str, sheet: str, cell: str) -> Dict[str, Any]:
    """Read one workbook cell, through the same reader the batch reads use.

    One implementation, so a single read and a batch read can never answer the
    same cell differently.
    """

    address = str(cell or "").upper()
    key: CellKey = (os.path.normcase(str(book_path or "")), str(sheet or ""), address)
    group = {
        "path": str(Path(str(book_path or "")).resolve()),
        "items": {key: {"sheet": str(sheet or ""), "cell": address}},
    }
    return _read_workbook_cells(group)[key]


def _group_cell_read_items(items: list) -> Tuple[Dict[str, Dict[str, Any]], List[CellKey]]:
    """Group requested cells by workbook, deduplicated, keeping caller order.

    The same workbook is opened once however many cells a caller asks it for,
    and every caller slot is answered from that one read. A linked range asks
    for thousands of cells of one workbook, so the path is resolved once per
    distinct path rather than once per cell: resolving touches the file system,
    and over a network share that alone cost seconds on a large range.
    """

    groups: Dict[str, Dict[str, Any]] = {}
    result_keys: List[CellKey] = []
    resolved_paths: Dict[str, str] = {}
    for item in items:
        raw_path = str(_item_field(item, "book_path") or "")
        resolved = resolved_paths.get(raw_path)
        if resolved is None:
            resolved = str(Path(raw_path).resolve())
            resolved_paths[raw_path] = resolved
        book_key = os.path.normcase(resolved)
        sheet = str(_item_field(item, "sheet") or "")
        cell = str(_item_field(item, "cell") or "").upper()
        cell_key = (book_key, sheet, cell)
        result_keys.append(cell_key)
        group = groups.setdefault(book_key, {"path": resolved, "items": {}})
        group["items"].setdefault(cell_key, {"sheet": sheet, "cell": cell})
    return groups, result_keys


def _cell_coordinate(cell: str) -> Tuple[int, int] | None:
    """Turn a cell address into its 1-based row and column, or None if it is not one."""

    try:
        row, column = coordinate_to_tuple(str(cell).replace("$", ""))
    except Exception:
        return None
    if row < 1 or column < 1:
        return None
    return row, column


def _read_sheet_cells(
    worksheet: Any,
    targets: List[Tuple[CellKey, Tuple[int, int]]],
) -> Dict[CellKey, Any]:
    """Answer every requested cell of one sheet from a single pass over it.

    A read-only worksheet re-parses the sheet XML from its first row every time
    it is asked for a cell by address, so a linked 120x120 range read cell by
    cell pays 14,400 parses of the same sheet and takes minutes. Walking the
    rectangle the requested cells span once answers all of them for the price
    of the single most expensive one.

    Cells the sheet does not reach are blank, exactly as a single address read
    of the same cell reports them.
    """

    wanted: Dict[Tuple[int, int], List[CellKey]] = {}
    for key, coordinate in targets:
        wanted.setdefault(coordinate, []).append(key)
    min_row = min(coordinate[0] for coordinate in wanted)
    max_row = max(coordinate[0] for coordinate in wanted)
    min_col = min(coordinate[1] for coordinate in wanted)
    max_col = max(coordinate[1] for coordinate in wanted)
    by_row: Dict[int, List[Tuple[int, List[CellKey]]]] = {}
    for (row, column), keys in wanted.items():
        by_row.setdefault(row, []).append((column, keys))

    values: Dict[CellKey, Any] = {key: None for key, _coordinate in targets}
    rows = worksheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=True,
    )
    # The walk is run to its end rather than abandoned once the last requested
    # cell has been answered: a read-only worksheet only lets go of the sheet's
    # XML stream when its walk finishes, and on Windows a stream left open
    # keeps the whole workbook file locked after the workbook has been closed.
    # The rectangle stops at the last requested row, so finishing costs nothing.
    for offset, row_values in enumerate(rows):
        columns = by_row.get(min_row + offset)
        if columns is None:
            continue
        for column, keys in columns:
            index = column - min_col
            raw = row_values[index] if index < len(row_values) else None
            for key in keys:
                values[key] = raw
    return values


def _read_workbook_cells(group: Dict[str, Any]) -> Dict[CellKey, Dict[str, Any]]:
    """Read every requested cell of one workbook in a single open.

    Each cell is answered on its own: an address the sheet cannot resolve, a
    missing sheet, or a non-numeric value is that cell's error and leaves the
    other cells of the same workbook with their real values, because callers
    validating a saved link need to know which reference broke.
    """

    book_path = str(group["path"])
    unique_items: Dict[CellKey, Dict[str, str]] = group["items"]
    workbook_results: Dict[CellKey, Dict[str, Any]] = {}
    p = Path(book_path)
    if not p.exists():
        return {
            key: {"ok": False, "error": f"File not found: {book_path}"}
            for key in unique_items
        }
    by_sheet: Dict[str, List[Tuple[CellKey, str]]] = {}
    for key, item in unique_items.items():
        by_sheet.setdefault(item["sheet"], []).append((key, item["cell"]))
    try:
        wb = openpyxl.load_workbook(str(p), data_only=True, read_only=True)
        try:
            for sheet, entries in by_sheet.items():
                if sheet not in wb.sheetnames:
                    for key, _cell in entries:
                        workbook_results[key] = {"ok": False, "error": f"Sheet not found: {sheet}"}
                    continue
                targets: List[Tuple[CellKey, Tuple[int, int]]] = []
                for key, cell in entries:
                    coordinate = _cell_coordinate(cell)
                    if coordinate is None:
                        workbook_results[key] = {
                            "ok": False,
                            "error": f"Cell not readable: {cell}",
                        }
                        continue
                    targets.append((key, coordinate))
                if not targets:
                    continue
                try:
                    raw_values = _read_sheet_cells(wb[sheet], targets)
                except Exception as sheet_error:
                    for key, _coordinate in targets:
                        workbook_results[key] = {
                            "ok": False,
                            "error": f"Sheet not readable: {sheet} ({sheet_error})",
                        }
                    continue
                for key, raw in raw_values.items():
                    workbook_results[key] = workbook_cell_value(raw)
        finally:
            wb.close()
    except Exception as e:
        for key in unique_items:
            workbook_results.setdefault(key, {"ok": False, "error": str(e)})
    return workbook_results


def _run_cell_read_batch(
    items: list,
    with_workbook_stats: bool,
    thread_name_prefix: str,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Run one grouped read pass and return the per-cell and per-workbook answers.

    When ``with_workbook_stats`` is set the worker that opens a workbook also
    stats it, so a caller that needs both the cell values and the file's
    modification time pays one pass over the (often network) share, not two.
    """

    groups, result_keys = _group_cell_read_items(items)

    def read_group(group: Dict[str, Any]) -> Tuple[Dict[CellKey, Dict[str, Any]], Dict[str, Any] | None]:
        cells = _read_workbook_cells(group)
        stats = _stat_workbook(str(group["path"])) if with_workbook_stats else None
        return cells, stats

    by_key: Dict[CellKey, Dict[str, Any]] = {}
    stats_by_key: Dict[str, Dict[str, Any]] = {}
    if groups:
        with ThreadPoolExecutor(
            max_workers=min(EXCEL_BATCH_MAX_WORKERS, len(groups)),
            thread_name_prefix=thread_name_prefix,
        ) as executor:
            futures = {
                executor.submit(read_group, group): book_key
                for book_key, group in groups.items()
            }
            for future, book_key in futures.items():
                cells, stats = future.result()
                by_key.update(cells)
                if stats is not None:
                    stats_by_key[book_key] = stats
    results = [dict(by_key[key]) for key in result_keys]
    workbooks = [dict(stats_by_key[book_key]) for book_key in groups if book_key in stats_by_key]
    return results, workbooks


def excel_read_cells_batch(items: list) -> Dict[str, Any]:
    results, _workbooks = _run_cell_read_batch(items, False, "arcrho-excel-check")
    return {"ok": True, "results": results}


def excel_validate_links(items: list) -> Dict[str, Any]:
    """Validate saved Excel link sources and report each workbook's timestamp.

    This is the check a dataset or DFM method runs when it opens: every stored
    reference is read where the app server can reach it, so a renamed sheet, a
    deleted row that left a ``#REF!``, or a workbook that moved is reported as
    that reference's own error rather than as a silent count. The workbook
    timestamps ride along from the same pass, so the caller can also tell
    whether an otherwise valid workbook is newer than the stored values
    without a second round trip over the share.
    """

    results, workbooks = _run_cell_read_batch(items, True, "arcrho-excel-validate")
    return {"ok": True, "results": results, "workbooks": workbooks}


def _run_workbook_path_batch(
    book_paths: list[str],
    read_one: Any,
    thread_name_prefix: str,
) -> Dict[str, Any]:
    """Run ``read_one`` once per distinct workbook path on a bounded pool.

    Callers pass the same path more than once - two datasets reading one
    workbook, a workbook listed under two aliases - so the paths are resolved
    and deduplicated before any I/O and the shared answer is copied back into
    every caller's slot. The pool keeps a listing off a network drive from
    paying one awaited round trip per file.
    """

    resolved_by_key: Dict[str, str] = {}
    result_keys: List[str] = []
    for index, raw_path in enumerate(book_paths):
        if not str(raw_path or "").strip():
            key = f"__invalid_path_{index}"
            result_keys.append(key)
            resolved_by_key[key] = ""
            continue
        resolved = str(Path(str(raw_path or "")).resolve())
        key = os.path.normcase(resolved)
        result_keys.append(key)
        resolved_by_key.setdefault(key, resolved)

    def read_item(item: tuple[str, str]) -> tuple[str, Dict[str, Any]]:
        key, resolved = item
        if not resolved:
            return key, {"ok": False, "path": resolved, "error": "Workbook path is empty."}
        return key, read_one(resolved)

    by_key: Dict[str, Dict[str, Any]] = {}
    if resolved_by_key:
        with ThreadPoolExecutor(
            max_workers=min(EXCEL_BATCH_MAX_WORKERS, len(resolved_by_key)),
            thread_name_prefix=thread_name_prefix,
        ) as executor:
            futures = [executor.submit(read_item, item) for item in resolved_by_key.items()]
            for future in futures:
                key, result = future.result()
                by_key[key] = result
    return {"ok": True, "results": [dict(by_key[key]) for key in result_keys]}


def _stat_workbook(resolved: str) -> Dict[str, Any]:
    try:
        stat_result = os.stat(resolved)
        if not stat.S_ISREG(stat_result.st_mode):
            return {"ok": False, "path": resolved, "error": f"File not found: {resolved}"}
        return {"ok": True, "path": resolved, "mtime": stat_result.st_mtime}
    except OSError as exc:
        return {"ok": False, "path": resolved, "error": str(exc)}


def _workbook_document_properties(resolved: str) -> Dict[str, str]:
    """Read the workbook's own Created/Modified/Last saved by properties.

    These are the document's record of itself - the same kind of answer a
    dataset sidecar gives about a dataset - rather than the file system's,
    so they survive a copy or a move that resets a file's creation time.
    They live in ``docProps/core.xml`` of the OOXML package, so a legacy
    ``.xls``, an encrypted workbook, and anything else that is not a readable
    zip simply have none; that is reported as blank, never as an error, because
    the workbook is still a perfectly good link target.
    """

    try:
        with zipfile.ZipFile(resolved) as package:
            with package.open(_CORE_PROPERTIES_ENTRY) as entry:
                raw = entry.read(_CORE_PROPERTIES_MAX_BYTES)
    except (KeyError, OSError, ValueError, zipfile.BadZipFile):
        return {}
    try:
        root = ElementTree.fromstring(raw)
    except ElementTree.ParseError:
        return {}

    def value(tag: str) -> str:
        found = root.find(tag, _CORE_PROPERTIES_NS)
        return str(found.text or "").strip() if found is not None else ""

    return {
        "created": value("dcterms:created"),
        "modified": value("dcterms:modified"),
        "last_modified_by": value("cp:lastModifiedBy"),
    }


def _stat_and_describe_workbook(resolved: str) -> Dict[str, Any]:
    result = _stat_workbook(resolved)
    if result.get("ok"):
        result.update(_workbook_document_properties(resolved))
    return result


def excel_file_mtimes_batch(book_paths: list[str]) -> Dict[str, Any]:
    return _run_workbook_path_batch(book_paths, _stat_workbook, "arcrho-excel-stat")


def excel_workbook_properties_batch(book_paths: list[str]) -> Dict[str, Any]:
    """``excel_file_mtimes_batch`` plus each workbook's document properties.

    One pass, not two: the stat and the small ``docProps`` read for a path
    happen in the same worker, so a listing over a network drive does not walk
    the same file list twice.
    """

    return _run_workbook_path_batch(
        book_paths, _stat_and_describe_workbook, "arcrho-excel-props"
    )


def excel_open_workbook(book_path: str, sheet: str = "", cell: str = "") -> Dict[str, Any]:
    p = Path(book_path).resolve()
    if not p.exists():
        return {"ok": False, "error": f"File not found: {book_path}"}
    try:
        import pythoncom
        import win32com.client
    except ImportError:
        return {"ok": False, "error": "win32com is not available on this system."}
    try:
        pythoncom.CoInitialize()
        try:
            xl = win32com.client.GetObject(Class="Excel.Application")
        except Exception:
            xl = win32com.client.Dispatch("Excel.Application")
            xl.Visible = True
        full_path = str(p)
        target_wb = None
        already_open = False
        for wb in xl.Workbooks:
            if wb.FullName.lower() == full_path.lower():
                target_wb = wb
                already_open = True
                break
        if target_wb is None:
            target_wb = xl.Workbooks.Open(str(p), ReadOnly=True)
        xl.Visible = True
        try:
            import win32gui
            hwnd = xl.Hwnd
            import win32con
            if win32gui.IsIconic(hwnd):
                win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
            win32gui.SetForegroundWindow(hwnd)
        except Exception:
            pass
        target_wb.Activate()
        if sheet and cell:
            try:
                ws = target_wb.Sheets(sheet)
                target = ws.Range(cell)
                scroll_row = max(1, target.Row - 10)
                scroll_col = max(1, target.Column - 10)
                ws.Activate()
                xl.Goto(ws.Cells(scroll_row, scroll_col), True)
                target.Select()
            except Exception:
                pass
        elif sheet:
            try:
                target_wb.Sheets(sheet).Activate()
            except Exception:
                pass
        return {"ok": True, "already_open": already_open}
    except Exception as e:
        return {"ok": False, "error": f"Failed to open workbook: {str(e)}"}
