"""Side-by-side ArcRho vs ResQ Result Selection dataset review workbook.

For every Result Selection (RS) method in the 17 reserving-class paths of
NJ_Annual_Prod_2026 Q3-Aug, this lays out the persisted ArcRho loaded-dataset
values and the live ResQ loaded-dataset values next to each other, with a
difference matrix (ArcRho minus ResQ) to their right -- one block per RS
method, grouped onto one sheet per reserving class. Only each dataset's own
input vector values are shown; weight selections are left out entirely, as
requested. A "Selected Ultimate" column is appended after the loaded
datasets, so an RS with n datasets loaded and m origin periods lays out as an
m*(n+1) matrix per side.

A Summary sheet links straight to every RS block that disagrees by more than
tolerance, is missing a loaded dataset on one side, or is missing from one
side entirely, so a reviewer does not have to hunt through 17 sheets to find
what needs attention.

Nothing is written back to ArcRho or ResQ.

Run with Python 3.10 from the repository root:

    py -3.10 python-api/migration/validation/rs_dataset_side_by_side_review.py
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import tempfile
from pathlib import Path
from typing import Any

_VALIDATION_DIR = Path(__file__).resolve().parent
_MIGRATION_DIR = _VALIDATION_DIR.parent
if str(_MIGRATION_DIR) not in sys.path:
    sys.path.insert(0, str(_MIGRATION_DIR))

import resq_data_migration as migration  # noqa: E402
from resq_migration.core import _encode_rc_folder, _normalize_import_name, _safe_attr  # noqa: E402


TARGET_PROJECT_NAME = "NJ_Annual_Prod_2026 Q3-Aug"
RC_PATHS = [
    r"PRNJ - PA\PA\NY\Direct Group\BI Total",
    r"PRNJ - PA\PA\NY\Direct Group\MP+PIP",
    r"PRNJ - PA\PA\Penn+CT\Direct Group\BI Total",
    r"PRNJ - PA\PA\Penn+CT\Direct Group\MP+PIP",
    r"PRNJ - PA\PA\All States\Direct Group\PD+UMPD",
    r"PRNJ - PA\PA\All States\Direct Group\COL",
    r"PRNJ - PA\PA\All States\Direct Group\CMPxCAT",
    r"PRNJ - PA\PA\NJ\Direct Group\MP+PIP",
    r"PRNJ - PA\PA\NJ\Direct Group\BIR51+UMBIR51",
    r"PRNJ - PA\PA\NJ\Direct Group\BIx51+UMBIx51",
    r"HPPREF\HO+DF\NJ\Legacy\HOL",
    r"HPPREF\HO+DF\NJ\Legacy\HOPxCAT",
    r"Rider\MC\All States\Direct Group\BI+PIP",
    r"Rider\MC\All States\Direct Group\PD+UMPD",
    r"Rider\MC\All States\Direct Group\PhysDxCat",
    r"PRNJ - PA\PA\MA\Direct Group\BI Total",
    r"PRNJ - PA\PA\MA\Direct Group\MP+PIP",
]

SELECTED_ULTIMATE_LABEL = "Selected Ultimate"

# RS values span wildly different magnitudes on one sheet -- claim counts in
# the tens next to incurred losses in the millions -- so a single absolute
# tolerance either misses real mismatches on small vectors or drowns in noise
# on large ones. Values are persisted at 6 decimal places, so agreement
# should be exact when both sides read the same ResQ data; anything past
# this combined floor is a genuine discrepancy worth a look.
ABS_TOLERANCE = 0.01
REL_TOLERANCE = 1e-4
NUMBER_FORMAT = "#,##0.0000"

_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


def _output_path(project_name: str) -> Path:
    return _VALIDATION_DIR / "results" / f"rs_dataset_side_by_side_{project_name}.xlsx"


def _tolerance(a_val: float, r_val: float) -> float:
    scale = max(abs(a_val), abs(r_val))
    return max(ABS_TOLERANCE, REL_TOLERANCE * scale)


def _read_arcrho_rs_methods(rc_dir: Path) -> dict[str, dict]:
    """Map RS name -> persisted ArcRho method payload for one reserving class."""

    methods_dir = rc_dir / "methods"
    out: dict[str, dict] = {}
    if not methods_dir.is_dir():
        return out
    for path in methods_dir.glob("RS@*.json"):
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        name = _normalize_import_name(payload.get("details_tab", {}).get("name")) or path.stem
        out[name] = payload
    return out


def _origin_labels(payload: dict) -> list[str]:
    labels = payload.get("method_tab", {}).get("origin_labels")
    return [str(label) for label in labels] if isinstance(labels, list) else []


def _dataset_columns(payload: dict) -> list[tuple[str, list[float | None]]]:
    """Ordered [(name, values)] for a payload's loaded datasets, Selected Ultimate last.

    Works for both the persisted ArcRho sidecar and the live ResQ payload
    returned by export_result_selection -- both share the same
    details_tab/method_tab shape, so one reader covers both sides.
    """

    method_tab = payload.get("method_tab", {})
    loaded = method_tab.get("loaded_datasets")
    columns: list[tuple[str, list[float | None]]] = []
    if isinstance(loaded, list):
        for dataset in loaded:
            if not isinstance(dataset, dict):
                continue
            # ResQ names sometimes carry stray runs of whitespace (e.g. "C 91 -
            # Current Qtr Indicated" with a double space) that a stale ArcRho
            # sidecar or a fresh COM read may or may not reproduce; normalize
            # both sides the same way so a whitespace quirk never masquerades
            # as a dataset present on one side only.
            name = _normalize_import_name(dataset.get("name")) or "(unnamed dataset)"
            values = dataset.get("values")
            columns.append((name, values if isinstance(values, list) else []))
    selected_ultimate = method_tab.get("selected_ultimate")
    columns.append((SELECTED_ULTIMATE_LABEL, selected_ultimate if isinstance(selected_ultimate, list) else []))
    return columns


def _build_rs_record(rc_path: str, name: str, arcrho_payload: dict | None, rs: Any) -> dict:
    arcrho_origin_labels = _origin_labels(arcrho_payload) if arcrho_payload else []
    arcrho_columns = _dataset_columns(arcrho_payload) if arcrho_payload else []

    resq_origin_labels: list[str] = []
    resq_columns: list[tuple[str, list[float | None]]] = []
    note_parts: list[str] = []

    if rs is not None:
        try:
            resq_payload = migration.export_result_selection(rs)
            resq_origin_labels = _origin_labels(resq_payload)
            resq_columns = _dataset_columns(resq_payload)
        except Exception as exc:
            note_parts.append(f"could not read ResQ result selection: {type(exc).__name__}: {exc}")
    else:
        note_parts.append("RS has a persisted ArcRho method JSON but was not found in ResQ")

    if arcrho_payload is None:
        note_parts.append("RS exists in ResQ but no persisted ArcRho method JSON was found")

    both_present = arcrho_payload is not None and bool(resq_columns)

    origin_count = max(len(arcrho_origin_labels), len(resq_origin_labels))
    origin_labels = list(arcrho_origin_labels or resq_origin_labels)
    for i in range(len(origin_labels), origin_count):
        origin_labels.append(str(i + 1))

    arcrho_by_name = {col_name: values for col_name, values in arcrho_columns}
    resq_by_name = {col_name: values for col_name, values in resq_columns}

    # Column order: every ArcRho dataset in ArcRho's own order, then any
    # ResQ-only dataset name not already present, with Selected Ultimate
    # always last regardless of which side introduced it.
    column_names = [n for n in arcrho_by_name if n != SELECTED_ULTIMATE_LABEL]
    for n in resq_by_name:
        if n != SELECTED_ULTIMATE_LABEL and n not in column_names:
            column_names.append(n)
    column_names.append(SELECTED_ULTIMATE_LABEL)

    if both_present:
        for col_name in column_names:
            if col_name == SELECTED_ULTIMATE_LABEL:
                continue
            if col_name not in arcrho_by_name:
                note_parts.append(f'dataset "{col_name}" loaded in ResQ only')
            elif col_name not in resq_by_name:
                note_parts.append(f'dataset "{col_name}" loaded in ArcRho only')

    arcrho_matrix: list[list[float | None]] = []
    resq_matrix: list[list[float | None]] = []
    diff_matrix: list[list[float | None]] = []
    max_abs_diff: float | None = None
    flagged_cells = 0
    only_arcrho_cells = 0
    only_resq_cells = 0

    for i in range(origin_count):
        a_row: list[float | None] = []
        r_row: list[float | None] = []
        d_row: list[float | None] = []
        for col_name in column_names:
            a_vals = arcrho_by_name.get(col_name) or []
            r_vals = resq_by_name.get(col_name) or []
            a_val = a_vals[i] if i < len(a_vals) else None
            r_val = r_vals[i] if i < len(r_vals) else None
            a_row.append(a_val)
            r_row.append(r_val)
            if a_val is None or r_val is None:
                d_row.append(None)
                # Only worth noting per-cell when the dataset is present on
                # both sides -- a whole missing dataset is already covered
                # by the note above.
                if both_present and col_name in arcrho_by_name and col_name in resq_by_name:
                    if a_val is not None:
                        only_arcrho_cells += 1
                    elif r_val is not None:
                        only_resq_cells += 1
                continue
            diff = float(a_val) - float(r_val)
            d_row.append(diff)
            if abs(diff) > _tolerance(float(a_val), float(r_val)):
                flagged_cells += 1
            if max_abs_diff is None or abs(diff) > max_abs_diff:
                max_abs_diff = abs(diff)
        arcrho_matrix.append(a_row)
        resq_matrix.append(r_row)
        diff_matrix.append(d_row)

    if only_arcrho_cells:
        note_parts.append(f"{only_arcrho_cells} cell(s) with a value present in ArcRho only")
    if only_resq_cells:
        note_parts.append(f"{only_resq_cells} cell(s) with a value present in ResQ only")

    note = "; ".join(note_parts)
    needs_review = bool(note) or (max_abs_diff is not None and max_abs_diff > ABS_TOLERANCE)

    return {
        "rc_path": rc_path,
        "name": name,
        "origin_labels": origin_labels,
        "column_names": column_names,
        "arcrho_matrix": arcrho_matrix,
        "resq_matrix": resq_matrix,
        "diff_matrix": diff_matrix,
        "origin_count": origin_count,
        "column_count": len(column_names),
        "max_abs_diff": max_abs_diff,
        "flagged_cells": flagged_cells,
        "note": note,
        "needs_review": needs_review,
    }


def run_comparison(
    *,
    project_name: str = TARGET_PROJECT_NAME,
    rc_paths: list[str] | None = None,
    app_factory=None,
    progress=print,
) -> tuple[list[dict], list[tuple[str, str]]]:
    """Compare every Result Selection method in scope.

    Returns (records, rc_errors) where records covers every RS found on
    either side and rc_errors lists reserving classes ResQ itself refused.
    """

    if app_factory is None:
        try:
            import win32com.client
        except ImportError as exc:
            raise RuntimeError("pywin32 is required: pip install pywin32") from exc

    rc_paths = list(rc_paths if rc_paths is not None else RC_PATHS)
    previous_scope = migration._apply_runtime_scope(project_name, migration.SERVER_ROOT)
    app = app_factory() if app_factory is not None else win32com.client.Dispatch("ResQ3Automation.ResQApplication")
    records: list[dict] = []
    rc_errors: list[tuple[str, str]] = []
    try:
        app.ConnectByName(migration.CONNECTION_NAME, migration.USER_NAME, migration.PASSWORD)
        project = app.Projects().Item(project_name)

        for rc_index, rc_path in enumerate(rc_paths, start=1):
            progress(f"RC {rc_index}/{len(rc_paths)}: {rc_path}")
            rc_dir = migration.PROJECT_DATA_DIR / _encode_rc_folder(rc_path)
            arcrho_methods = _read_arcrho_rs_methods(rc_dir)

            try:
                reserving_class = project.ReservingClasses().Item(rc_path)
                rs_collection = list(reserving_class.ResultSelections())
            except Exception as exc:
                rc_errors.append((rc_path, f"could not read ResQ reserving class: {type(exc).__name__}: {exc}"))
                continue

            resq_rs: dict[str, Any] = {}
            for rs in rs_collection:
                name = _normalize_import_name(_safe_attr(rs, "Name", ""))
                if not name:
                    continue
                resq_rs[name] = rs

            all_names = sorted(set(arcrho_methods) | set(resq_rs), key=str.casefold)
            for name in all_names:
                records.append(
                    _build_rs_record(rc_path, name, arcrho_methods.get(name), resq_rs.get(name))
                )
    finally:
        try:
            app.Disconnect()
        except Exception:
            pass
        migration._restore_runtime_scope(previous_scope)
    return records, rc_errors


def _sheet_title(rc_path: str, used: set[str]) -> str:
    segments = [segment.strip() for segment in rc_path.split("\\") if segment.strip()]
    label = f"{segments[-3]} {segments[-1]}" if len(segments) >= 3 else rc_path
    label = _INVALID_SHEET_CHARS.sub("-", label)[:31]
    base = label
    suffix_index = 2
    while label.casefold() in used:
        suffix = f" ({suffix_index})"
        label = base[: 31 - len(suffix)] + suffix
        suffix_index += 1
    used.add(label.casefold())
    return label


def _write_rs_block(sheet, start_row: int, record: dict, styles: dict) -> tuple[int, int]:
    """Write one RS's ArcRho | ResQ | Diff dataset matrices. Returns (header_row, next_free_row)."""

    column_count = max(record["column_count"], 1)
    origin_count = record["origin_count"]
    gap = 1

    header_row = start_row
    header_cell = sheet.cell(row=header_row, column=1, value=record["name"])
    header_cell.font = styles["bold"]
    if record["note"]:
        note_cell = sheet.cell(row=header_row, column=2, value=record["note"])
        note_cell.font = styles["note"]

    group_row = header_row + 1
    label_row = header_row + 2
    data_start_row = header_row + 3

    origin_col = 1
    arcrho_start_col = 2
    resq_start_col = arcrho_start_col + column_count + gap
    diff_start_col = resq_start_col + column_count + gap

    def _group_header(col: int, text: str) -> None:
        cell = sheet.cell(row=group_row, column=col, value=text)
        cell.font = styles["bold"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center"]
        if column_count > 1:
            sheet.merge_cells(start_row=group_row, start_column=col, end_row=group_row, end_column=col + column_count - 1)

    _group_header(arcrho_start_col, "ArcRho")
    _group_header(resq_start_col, "ResQ")
    _group_header(diff_start_col, "Diff (ArcRho − ResQ)")

    sheet.cell(row=label_row, column=origin_col, value="Origin").font = styles["bold"]
    for j, col_name in enumerate(record["column_names"]):
        is_ultimate = col_name == SELECTED_ULTIMATE_LABEL
        for base_col in (arcrho_start_col, resq_start_col, diff_start_col):
            cell = sheet.cell(row=label_row, column=base_col + j, value=col_name)
            cell.font = styles["bold"]
            if is_ultimate:
                cell.fill = styles["ultimate_fill"]

    for i in range(origin_count):
        row = data_start_row + i
        origin_label = record["origin_labels"][i] if i < len(record["origin_labels"]) else i + 1
        sheet.cell(row=row, column=origin_col, value=origin_label)
        a_row = record["arcrho_matrix"][i] if i < len(record["arcrho_matrix"]) else []
        r_row = record["resq_matrix"][i] if i < len(record["resq_matrix"]) else []
        d_row = record["diff_matrix"][i] if i < len(record["diff_matrix"]) else []
        for j in range(column_count):
            a_val = a_row[j] if j < len(a_row) else None
            r_val = r_row[j] if j < len(r_row) else None
            d_val = d_row[j] if j < len(d_row) else None
            flagged = (
                d_val is not None
                and a_val is not None
                and r_val is not None
                and abs(d_val) > _tolerance(float(a_val), float(r_val))
            )

            a_cell = sheet.cell(row=row, column=arcrho_start_col + j, value=a_val)
            r_cell = sheet.cell(row=row, column=resq_start_col + j, value=r_val)
            d_cell = sheet.cell(row=row, column=diff_start_col + j, value=d_val)
            for cell in (a_cell, r_cell, d_cell):
                cell.number_format = NUMBER_FORMAT
                if flagged:
                    cell.fill = styles["flag_fill"]

    next_row = data_start_row + max(origin_count, 1)
    return header_row, next_row


def _autosize(sheet, *, min_width: int, max_width: int) -> None:
    # A merged group header ("ArcRho" / "ResQ" / "Diff (ArcRho − ResQ)") is
    # anchored on one cell but reads across every column of that RS's block.
    # Different RS methods on the same sheet merge different spans, so over
    # many blocks nearly every column ends up anchoring some block's header
    # at some point; counting that text toward a single column's width would
    # balloon every dataset column to fit "Diff (ArcRho − ResQ)".
    wide_merge_anchors = {
        (merged_range.min_row, merged_range.min_col)
        for merged_range in sheet.merged_cells.ranges
        if merged_range.max_col > merged_range.min_col
    }
    for column_cells in sheet.columns:
        column_letter = None
        width = 0
        for cell in column_cells:
            letter = getattr(cell, "column_letter", None)
            if letter is None:
                continue
            column_letter = letter
            if (cell.row, cell.column) in wide_merge_anchors:
                continue
            text = str(cell.value) if cell.value is not None else ""
            width = max(width, len(text))
        if column_letter:
            sheet.column_dimensions[column_letter].width = min(max(width + 2, min_width), max_width)


def write_workbook(
    path: Path,
    records: list[dict],
    rc_errors: list[tuple[str, str]],
    *,
    project_name: str = TARGET_PROJECT_NAME,
    rc_paths: list[str] | None = None,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    rc_paths = list(rc_paths if rc_paths is not None else RC_PATHS)

    styles = {
        "bold": Font(bold=True),
        "note": Font(italic=True, color="9C0006"),
        "header_fill": PatternFill("solid", fgColor="DDEBF7"),
        "ultimate_fill": PatternFill("solid", fgColor="FFF2CC"),
        "flag_fill": PatternFill("solid", fgColor="FFC7CE"),
        "center": Alignment(horizontal="center"),
        "link": Font(color="0563C1", underline="single"),
    }

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    used_titles = {"summary"}
    sheet_by_rc = {rc_path: workbook.create_sheet(_sheet_title(rc_path, used_titles)) for rc_path in rc_paths}

    records_by_rc: dict[str, list[dict]] = {}
    for record in records:
        records_by_rc.setdefault(record["rc_path"], []).append(record)

    rs_anchor: dict[tuple[str, str], tuple[str, int]] = {}
    for rc_path in rc_paths:
        sheet = sheet_by_rc[rc_path]
        title_cell = sheet.cell(row=1, column=1, value=rc_path)
        title_cell.font = styles["bold"]
        row = 3
        for record in sorted(records_by_rc.get(rc_path, []), key=lambda r: r["name"].casefold()):
            header_row, next_row = _write_rs_block(sheet, row, record, styles)
            rs_anchor[(rc_path, record["name"])] = (sheet.title, header_row)
            row = next_row + 2
        _autosize(sheet, min_width=9, max_width=22)
        sheet.column_dimensions["A"].width = 26
        sheet.freeze_panes = "A3"

    summary_sheet.cell(
        row=1,
        column=1,
        value=(
            f"Project: {project_name}    Tolerance: max({ABS_TOLERANCE:g}, {REL_TOLERANCE:g} * |value|)"
            "    weight selections are excluded, only loaded-dataset values and Selected Ultimate are shown"
        ),
    ).font = styles["bold"]
    header_row_index = 3
    headers = ["RC Path", "RS Name", "Max Abs Diff", "Flagged Cells", "Note"]
    for col, text in enumerate(headers, start=1):
        cell = summary_sheet.cell(row=header_row_index, column=col, value=text)
        cell.font = styles["bold"]
    summary_sheet.freeze_panes = f"A{header_row_index + 1}"

    review_records = [record for record in records if record["needs_review"]]
    row = header_row_index + 1
    for record in sorted(review_records, key=lambda r: (r["rc_path"], r["name"].casefold())):
        summary_sheet.cell(row=row, column=1, value=record["rc_path"])
        name_cell = summary_sheet.cell(row=row, column=2, value=record["name"])
        anchor = rs_anchor.get((record["rc_path"], record["name"]))
        if anchor:
            anchor_sheet_title, anchor_row = anchor
            name_cell.hyperlink = f"#'{anchor_sheet_title}'!A{anchor_row}"
            name_cell.font = styles["link"]
        if record["max_abs_diff"] is not None:
            diff_cell = summary_sheet.cell(row=row, column=3, value=record["max_abs_diff"])
            diff_cell.number_format = NUMBER_FORMAT
        summary_sheet.cell(row=row, column=4, value=record["flagged_cells"] or None)
        summary_sheet.cell(row=row, column=5, value=record["note"])
        row += 1

    for rc_path, note in rc_errors:
        summary_sheet.cell(row=row, column=1, value=rc_path)
        summary_sheet.cell(row=row, column=2, value="(reserving class)")
        summary_sheet.cell(row=row, column=5, value=note)
        row += 1

    if row == header_row_index + 1:
        summary_sheet.cell(row=row, column=1, value="No RS matrices need review.")

    _autosize(summary_sheet, min_width=12, max_width=80)

    descriptor, temporary_name = tempfile.mkstemp(prefix=f".{path.stem}-", suffix=".xlsx", dir=path.parent)
    os.close(descriptor)
    temporary_path = Path(temporary_name)
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, path)
    finally:
        workbook.close()
        if temporary_path.exists():
            temporary_path.unlink()


def _parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--project", default=TARGET_PROJECT_NAME, help="ResQ project name to review.")
    parser.add_argument(
        "--rc",
        action="append",
        help="Only review reserving classes whose path contains this text; repeatable.",
    )
    parser.add_argument("--no-open", action="store_true", help="Do not open the workbook when the run finishes.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = _parse_args(argv)
    rc_paths = RC_PATHS
    if args.rc:
        needles = [text.casefold() for text in args.rc]
        rc_paths = [path for path in RC_PATHS if any(needle in path.casefold() for needle in needles)]
        if not rc_paths:
            print("No reserving class matched --rc.")
            return 2

    records, rc_errors = run_comparison(project_name=args.project, rc_paths=rc_paths)
    output_path = _output_path(args.project)
    write_workbook(output_path, records, rc_errors, project_name=args.project, rc_paths=rc_paths)
    review_records = [record for record in records if record["needs_review"]]
    needs_attention = bool(review_records) or bool(rc_errors)
    print(f"Compared {len(rc_paths)} reserving class(es), {len(records)} Result Selection method(s).")
    print(f"{len(review_records)} RS(s) need review" + (f", {len(rc_errors)} reserving class(es) could not be read" if rc_errors else "") + ".")
    print(f"Excel report: {output_path}")
    if needs_attention and not args.no_open:
        try:
            os.startfile(output_path)  # noqa: S606 - opening the report just written, for the operator running this script
        except Exception as exc:
            print(f"Could not open the report automatically: {type(exc).__name__}: {exc}")
    return 0 if not needs_attention else 1


if __name__ == "__main__":
    raise SystemExit(main())
