"""Side-by-side ArcRho vs ResQ DFM ratio triangle review workbook.

For every DFM in the 17 reserving-class paths of NJ_Annual_Prod_2026 Q3-Aug,
this lays the persisted ArcRho ratio triangle and the live ResQ ratio
triangle next to each other, with a difference triangle (ArcRho minus ResQ)
to their right, followed by the DFM's output (ultimate) vector compared the
same way -- one sheet per reserving class. A Summary sheet links straight to
every DFM that disagrees on a ratio cell or an ultimate value by more than
tolerance, or that is missing from one side, so a reviewer does not have to
hunt through 17 sheets to find what needs attention.

DFMs whose name contains "Adjusted" are left out entirely: their ratios are
deliberately altered from the raw ResQ triangle by the reserve review's
growth/cutoff adjustment and are not expected to match ResQ.

Nothing is written back to ArcRho or ResQ.

Run with Python 3.10 from the repository root:

    py -3.10 python-api/migration/validation/dfm_ratio_side_by_side_review.py
"""

from __future__ import annotations

import json
import os
import re
import sys
import tempfile
from pathlib import Path
from typing import Any

_VALIDATION_DIR = Path(__file__).resolve().parent
_MIGRATION_DIR = _VALIDATION_DIR.parent
if str(_MIGRATION_DIR) not in sys.path:
    sys.path.insert(0, str(_MIGRATION_DIR))

import resq_data_migration as migration  # noqa: E402
from resq_migration.core import _clean_name, _encode_rc_folder, _safe_attr  # noqa: E402


TARGET_PROJECT_NAME = "NJ_Annual_Prod_2026 Q3-Aug"
RC_PATHS = [
    r"PRNJ - PA\PA\NY\Direct Group\BI Total",
    r"PRNJ - PA\PA\NY\Direct Group\MP+PIP",
    r"PRNJ - PA\PA\Penn+CT\Direct Group\BI Total",
    r"PRNJ - PA\PA\Penn+CT\Direct Group\MP+PIP",
    r"PRNJ - PA\PA\All States\Direct Group\PD+UMPD",
    r"PRNJ - PA\PA\All States\Direct Group\COL",
    r"PRNJ - PA\PA\All States\Direct Group\CMPxCAT",
    r"PRNJ - PA\PA\NJ\Direct Group\MP+PIP",
    r"PRNJ - PA\PA\NJ\Direct Group\BIR51+UMBIR51",
    r"PRNJ - PA\PA\NJ\Direct Group\BIx51+UMBIx51",
    r"HPPREF\HO+DF\NJ\Legacy\HOL",
    r"HPPREF\HO+DF\NJ\Legacy\HOPxCAT",
    r"Rider\MC\All States\Direct Group\BI+PIP",
    r"Rider\MC\All States\Direct Group\PD+UMPD",
    r"Rider\MC\All States\Direct Group\PhysDxCat",
    r"PRNJ - PA\PA\MA\Direct Group\BI Total",
    r"PRNJ - PA\PA\MA\Direct Group\MP+PIP",
]

DECIMAL_PLACES = 4
TOLERANCE = 0.5 * 10 ** (-DECIMAL_PLACES)
NUMBER_FORMAT = "0.0000"

# Ultimate (output) vector values are dollar amounts, not ratios, so they are
# compared at two decimal places -- the same precision the ResQ import uses
# when it cross-checks an Engine-built dataset against ResQ.
ULTIMATE_DECIMAL_PLACES = 2
ULTIMATE_TOLERANCE = 0.5 * 10 ** (-ULTIMATE_DECIMAL_PLACES)
ULTIMATE_NUMBER_FORMAT = "#,##0.00"

OUTPUT_PATH = _VALIDATION_DIR / "results" / f"dfm_ratio_side_by_side_{TARGET_PROJECT_NAME}.xlsx"

_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


def _is_adjusted(name: str) -> bool:
    return "adjusted" in name.lower()


def _read_arcrho_dfm_methods(rc_dir: Path) -> dict[str, dict]:
    """Map DFM name -> persisted ArcRho method payload for one reserving class."""

    methods_dir = rc_dir / "methods"
    out: dict[str, dict] = {}
    if not methods_dir.is_dir():
        return out
    for path in methods_dir.glob("DFM@*.json"):
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        name = _clean_name(payload.get("details_tab", {}).get("name")) or path.stem
        if _is_adjusted(name):
            continue
        out[name] = payload
    return out


def _arcrho_ratio_matrix(payload: dict) -> list[list[float | None]]:
    ratio_triangle = payload.get("ratios_tab", {}).get("ratio_triangle", {})
    values = ratio_triangle.get("ratio_values")
    return values if isinstance(values, list) else []


def _arcrho_dev_labels(payload: dict) -> list[str]:
    ratio_triangle = payload.get("ratios_tab", {}).get("ratio_triangle", {})
    labels = ratio_triangle.get("development_labels")
    return [str(label) for label in labels] if isinstance(labels, list) else []


def _arcrho_origin_labels(payload: dict) -> list[str]:
    ratio_triangle = payload.get("ratios_tab", {}).get("ratio_triangle", {})
    labels = ratio_triangle.get("origin_labels") or payload.get("data_tab", {}).get("origin_labels")
    return [str(label) for label in labels] if isinstance(labels, list) else []


def _arcrho_ultimate_vector(payload: dict) -> list[float | None]:
    values = payload.get("results_tab", {}).get("ultimate_vector")
    return [float(v) if isinstance(v, (int, float)) else None for v in values] if isinstance(values, list) else []


def _resq_ultimate_vector(dfm, origin_count: int) -> list[float | None]:
    values: list[float | None] = []
    for i in range(1, origin_count + 1):
        try:
            v = dfm.Ultimates(i)
            values.append(float(v) if v is not None else None)
        except Exception:
            values.append(None)
    return values


# ResQ's ExcludedRatios code for a cell with no defined ratio (typically a 0/0
# development step). It reports a 1.0 placeholder value there, not a real
# ratio, so it must be read as blank -- the same way ArcRho's own ratio
# triangle omits it -- rather than compared against ArcRho's missing cell.
RESQ_RATIO_EMPTY_CELL = 2


def _resq_ratio_matrix(dfm) -> list[list[float | None]]:
    origin_count = int(_safe_attr(dfm, "OriginCount", 0) or 0)
    matrix: list[list[float | None]] = []
    for i in range(1, origin_count + 1):
        try:
            row_dev = int(dfm.DevelopmentCount(i))
        except Exception:
            row_dev = 0
        row: list[float | None] = []
        for j in range(1, row_dev + 1):
            try:
                excluded_code = int(dfm.ExcludedRatios(i, j))
            except Exception:
                excluded_code = 0
            if excluded_code == RESQ_RATIO_EMPTY_CELL:
                row.append(None)
                continue
            try:
                v = dfm.Ratios(OriginIndex=i, DevIndex=j)
                row.append(float(v) if v is not None else None)
            except Exception:
                row.append(None)
        matrix.append(row)
    return matrix


def _resq_dev_labels(dfm, max_dev: int) -> list[str]:
    labels: list[str] = []
    for j in range(1, max_dev + 1):
        try:
            labels.append(_clean_name(dfm.DevelopmentLabel(j)))
        except Exception:
            labels.append("")
    return labels


def _resq_origin_labels(dfm, origin_count: int) -> list[str]:
    triangle = _safe_attr(dfm, "InputTriangle", None)
    labels: list[str] = []
    for i in range(1, origin_count + 1):
        label = None
        if triangle is not None:
            try:
                label = triangle.OriginLabel(i)
            except Exception:
                label = None
        labels.append(_clean_name(label) if label else str(i))
    return labels


def _build_dfm_record(rc_path: str, name: str, arcrho_payload: dict | None, dfm: Any) -> dict:
    arcrho_matrix = _arcrho_ratio_matrix(arcrho_payload) if arcrho_payload else []
    arcrho_dev_labels = _arcrho_dev_labels(arcrho_payload) if arcrho_payload else []
    arcrho_origin_labels = _arcrho_origin_labels(arcrho_payload) if arcrho_payload else []

    arcrho_ultimate = _arcrho_ultimate_vector(arcrho_payload) if arcrho_payload else []

    resq_matrix: list[list[float | None]] = []
    resq_dev_labels: list[str] = []
    resq_origin_labels: list[str] = []
    resq_ultimate: list[float | None] = []
    note_parts: list[str] = []

    if dfm is not None:
        try:
            resq_matrix = _resq_ratio_matrix(dfm)
        except Exception as exc:
            note_parts.append(f"could not read ResQ ratios: {type(exc).__name__}: {exc}")
        resq_origin_count = len(resq_matrix)
        max_dev = max((len(row) for row in resq_matrix), default=0)
        resq_dev_labels = _resq_dev_labels(dfm, max_dev)
        resq_origin_labels = _resq_origin_labels(dfm, resq_origin_count)
        try:
            resq_ultimate = _resq_ultimate_vector(dfm, max(resq_origin_count, len(arcrho_ultimate)))
        except Exception as exc:
            note_parts.append(f"could not read ResQ ultimate values: {type(exc).__name__}: {exc}")
    else:
        note_parts.append("DFM has a persisted ArcRho method JSON but was not found in ResQ")

    if arcrho_payload is None:
        note_parts.append("DFM exists in ResQ but no persisted ArcRho method JSON was found")

    origin_count = max(len(arcrho_matrix), len(resq_matrix), len(arcrho_ultimate), len(resq_ultimate))
    dev_count = max(
        max((len(r) for r in arcrho_matrix), default=0),
        max((len(r) for r in resq_matrix), default=0),
    )

    origin_labels = list(arcrho_origin_labels or resq_origin_labels)
    for i in range(len(origin_labels), origin_count):
        origin_labels.append(str(i + 1))
    arcrho_dev_labels_out = list(arcrho_dev_labels or resq_dev_labels)
    resq_dev_labels_out = list(resq_dev_labels or arcrho_dev_labels)
    for j in range(len(arcrho_dev_labels_out), dev_count):
        arcrho_dev_labels_out.append(f"dev {j + 1}")
    for j in range(len(resq_dev_labels_out), dev_count):
        resq_dev_labels_out.append(f"dev {j + 1}")

    diff_matrix: list[list[float | None]] = []
    max_abs_diff: float | None = None
    flagged_cells = 0
    only_arcrho_cells = 0
    only_resq_cells = 0
    for i in range(origin_count):
        a_row = arcrho_matrix[i] if i < len(arcrho_matrix) else []
        r_row = resq_matrix[i] if i < len(resq_matrix) else []
        diff_row: list[float | None] = []
        for j in range(dev_count):
            a_val = a_row[j] if j < len(a_row) else None
            r_val = r_row[j] if j < len(r_row) else None
            if a_val is None or r_val is None:
                diff_row.append(None)
                # Only worth noting when the DFM is genuinely present on both
                # sides -- otherwise every cell is "only" on the one side that
                # has the DFM at all, which is already covered by note_parts.
                if arcrho_payload is not None and dfm is not None:
                    if a_val is not None:
                        only_arcrho_cells += 1
                    elif r_val is not None:
                        only_resq_cells += 1
                continue
            diff = float(a_val) - float(r_val)
            diff_row.append(diff)
            if abs(diff) > TOLERANCE:
                flagged_cells += 1
            if max_abs_diff is None or abs(diff) > max_abs_diff:
                max_abs_diff = abs(diff)
        diff_matrix.append(diff_row)

    if only_arcrho_cells:
        note_parts.append(f"{only_arcrho_cells} cell(s) with a ratio computed in ArcRho only")
    if only_resq_cells:
        note_parts.append(f"{only_resq_cells} cell(s) with a ratio computed in ResQ only")

    ultimate_diff: list[float | None] = []
    ultimate_max_abs_diff: float | None = None
    ultimate_flagged_cells = 0
    only_arcrho_ultimate = 0
    only_resq_ultimate = 0
    for i in range(origin_count):
        a_val = arcrho_ultimate[i] if i < len(arcrho_ultimate) else None
        r_val = resq_ultimate[i] if i < len(resq_ultimate) else None
        if a_val is None or r_val is None:
            ultimate_diff.append(None)
            if arcrho_payload is not None and dfm is not None:
                if a_val is not None:
                    only_arcrho_ultimate += 1
                elif r_val is not None:
                    only_resq_ultimate += 1
            continue
        diff = float(a_val) - float(r_val)
        ultimate_diff.append(diff)
        if abs(diff) > ULTIMATE_TOLERANCE:
            ultimate_flagged_cells += 1
        if ultimate_max_abs_diff is None or abs(diff) > ultimate_max_abs_diff:
            ultimate_max_abs_diff = abs(diff)

    if only_arcrho_ultimate:
        note_parts.append(f"{only_arcrho_ultimate} ultimate value(s) computed in ArcRho only")
    if only_resq_ultimate:
        note_parts.append(f"{only_resq_ultimate} ultimate value(s) computed in ResQ only")

    note = "; ".join(note_parts)
    needs_review = (
        bool(note)
        or (max_abs_diff is not None and max_abs_diff > TOLERANCE)
        or (ultimate_max_abs_diff is not None and ultimate_max_abs_diff > ULTIMATE_TOLERANCE)
    )

    return {
        "rc_path": rc_path,
        "name": name,
        "origin_labels": origin_labels,
        "arcrho_dev_labels": arcrho_dev_labels_out,
        "resq_dev_labels": resq_dev_labels_out,
        "arcrho_matrix": arcrho_matrix,
        "resq_matrix": resq_matrix,
        "arcrho_ultimate": arcrho_ultimate,
        "resq_ultimate": resq_ultimate,
        "ultimate_diff": ultimate_diff,
        "ultimate_max_abs_diff": ultimate_max_abs_diff,
        "ultimate_flagged_cells": ultimate_flagged_cells,
        "diff_matrix": diff_matrix,
        "origin_count": origin_count,
        "dev_count": dev_count,
        "max_abs_diff": max_abs_diff,
        "flagged_cells": flagged_cells,
        "note": note,
        "needs_review": needs_review,
    }


def run_comparison(app_factory=None, progress=print) -> tuple[list[dict], list[tuple[str, str]]]:
    """Compare every non-adjusted DFM in scope.

    Returns (records, rc_errors) where records covers every DFM found on
    either side and rc_errors lists reserving classes ResQ itself refused.
    """

    try:
        import win32com.client
    except ImportError as exc:
        raise RuntimeError("pywin32 is required: pip install pywin32") from exc

    previous_scope = migration._apply_runtime_scope(TARGET_PROJECT_NAME, migration.SERVER_ROOT)
    app = app_factory() if app_factory is not None else win32com.client.Dispatch("ResQ3Automation.ResQApplication")
    records: list[dict] = []
    rc_errors: list[tuple[str, str]] = []
    try:
        app.ConnectByName(migration.CONNECTION_NAME, migration.USER_NAME, migration.PASSWORD)
        project = app.Projects().Item(TARGET_PROJECT_NAME)

        for rc_index, rc_path in enumerate(RC_PATHS, start=1):
            progress(f"RC {rc_index}/{len(RC_PATHS)}: {rc_path}")
            rc_dir = migration.PROJECT_DATA_DIR / _encode_rc_folder(rc_path)
            arcrho_methods = _read_arcrho_dfm_methods(rc_dir)

            try:
                reserving_class = project.ReservingClasses().Item(rc_path)
                dfm_collection = list(reserving_class.DFMMethods())
            except Exception as exc:
                rc_errors.append((rc_path, f"could not read ResQ reserving class: {type(exc).__name__}: {exc}"))
                continue

            resq_dfms: dict[str, Any] = {}
            for dfm in dfm_collection:
                name = _clean_name(_safe_attr(dfm, "Name", ""))
                if not name or _is_adjusted(name):
                    continue
                resq_dfms[name] = dfm

            all_names = sorted(set(arcrho_methods) | set(resq_dfms), key=str.casefold)
            for name in all_names:
                records.append(
                    _build_dfm_record(rc_path, name, arcrho_methods.get(name), resq_dfms.get(name))
                )
    finally:
        try:
            app.Disconnect()
        except Exception:
            pass
        migration._restore_runtime_scope(previous_scope)
    return records, rc_errors


def _sheet_title(rc_path: str, used: set[str]) -> str:
    segments = [segment.strip() for segment in rc_path.split("\\") if segment.strip()]
    label = f"{segments[-3]} {segments[-1]}" if len(segments) >= 3 else rc_path
    label = _INVALID_SHEET_CHARS.sub("-", label)[:31]
    base = label
    suffix_index = 2
    while label.casefold() in used:
        suffix = f" ({suffix_index})"
        label = base[: 31 - len(suffix)] + suffix
        suffix_index += 1
    used.add(label.casefold())
    return label


def _write_dfm_block(sheet, start_row: int, record: dict, styles: dict) -> tuple[int, int]:
    """Write one DFM's ArcRho | ResQ | Diff triangles. Returns (header_row, next_free_row)."""

    dev_count = max(record["dev_count"], 1)
    origin_count = record["origin_count"]
    gap = 1

    header_row = start_row
    header_cell = sheet.cell(row=header_row, column=1, value=record["name"])
    header_cell.font = styles["bold"]
    if record["note"]:
        note_cell = sheet.cell(row=header_row, column=2, value=record["note"])
        note_cell.font = styles["note"]

    group_row = header_row + 1
    label_row = header_row + 2
    data_start_row = header_row + 3

    origin_col = 1
    arcrho_start_col = 2
    resq_start_col = arcrho_start_col + dev_count + gap
    diff_start_col = resq_start_col + dev_count + gap
    ultimate_start_col = diff_start_col + dev_count + gap + gap

    def _group_header(col: int, text: str, width: int) -> None:
        cell = sheet.cell(row=group_row, column=col, value=text)
        cell.font = styles["bold"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center"]
        if width > 1:
            sheet.merge_cells(start_row=group_row, start_column=col, end_row=group_row, end_column=col + width - 1)

    _group_header(arcrho_start_col, "ArcRho", dev_count)
    _group_header(resq_start_col, "ResQ", dev_count)
    _group_header(diff_start_col, "Diff (ArcRho − ResQ)", dev_count)
    _group_header(ultimate_start_col, "Output vector (ultimate)", 3)

    sheet.cell(row=label_row, column=origin_col, value="Origin").font = styles["bold"]
    for j in range(dev_count):
        a_label = record["arcrho_dev_labels"][j] if j < len(record["arcrho_dev_labels"]) else ""
        r_label = record["resq_dev_labels"][j] if j < len(record["resq_dev_labels"]) else ""
        sheet.cell(row=label_row, column=arcrho_start_col + j, value=a_label).font = styles["bold"]
        sheet.cell(row=label_row, column=resq_start_col + j, value=r_label).font = styles["bold"]
        sheet.cell(row=label_row, column=diff_start_col + j, value=a_label or r_label).font = styles["bold"]
    for j, text in enumerate(("ArcRho", "ResQ", "Diff")):
        sheet.cell(row=label_row, column=ultimate_start_col + j, value=text).font = styles["bold"]

    for i in range(origin_count):
        row = data_start_row + i
        origin_label = record["origin_labels"][i] if i < len(record["origin_labels"]) else i + 1
        sheet.cell(row=row, column=origin_col, value=origin_label)
        a_row = record["arcrho_matrix"][i] if i < len(record["arcrho_matrix"]) else []
        r_row = record["resq_matrix"][i] if i < len(record["resq_matrix"]) else []
        d_row = record["diff_matrix"][i] if i < len(record["diff_matrix"]) else []
        for j in range(dev_count):
            a_val = a_row[j] if j < len(a_row) else None
            r_val = r_row[j] if j < len(r_row) else None
            d_val = d_row[j] if j < len(d_row) else None
            flagged = d_val is not None and abs(d_val) > TOLERANCE

            a_cell = sheet.cell(row=row, column=arcrho_start_col + j, value=a_val)
            r_cell = sheet.cell(row=row, column=resq_start_col + j, value=r_val)
            d_cell = sheet.cell(row=row, column=diff_start_col + j, value=d_val)
            for cell in (a_cell, r_cell, d_cell):
                cell.number_format = NUMBER_FORMAT
                if flagged:
                    cell.fill = styles["flag_fill"]

        au_val = record["arcrho_ultimate"][i] if i < len(record["arcrho_ultimate"]) else None
        ru_val = record["resq_ultimate"][i] if i < len(record["resq_ultimate"]) else None
        du_val = record["ultimate_diff"][i] if i < len(record["ultimate_diff"]) else None
        ultimate_flagged = du_val is not None and abs(du_val) > ULTIMATE_TOLERANCE

        au_cell = sheet.cell(row=row, column=ultimate_start_col, value=au_val)
        ru_cell = sheet.cell(row=row, column=ultimate_start_col + 1, value=ru_val)
        du_cell = sheet.cell(row=row, column=ultimate_start_col + 2, value=du_val)
        for cell in (au_cell, ru_cell, du_cell):
            cell.number_format = ULTIMATE_NUMBER_FORMAT
            if ultimate_flagged:
                cell.fill = styles["flag_fill"]

    next_row = data_start_row + max(origin_count, 1)
    return header_row, next_row


def _autosize(sheet, *, min_width: int, max_width: int) -> None:
    # A merged group header ("ArcRho" / "ResQ" / "Diff (ArcRho − ResQ)") is
    # anchored on one cell but reads across every column of that DFM's block.
    # Different DFMs on the same sheet merge different spans, so over many
    # blocks nearly every column ends up anchoring some block's header at
    # some point; counting that text toward a single column's width would
    # balloon every ratio column to fit "Diff (ArcRho − ResQ)".
    wide_merge_anchors = {
        (merged_range.min_row, merged_range.min_col)
        for merged_range in sheet.merged_cells.ranges
        if merged_range.max_col > merged_range.min_col
    }
    for column_cells in sheet.columns:
        column_letter = None
        width = 0
        for cell in column_cells:
            letter = getattr(cell, "column_letter", None)
            if letter is None:
                continue
            column_letter = letter
            if (cell.row, cell.column) in wide_merge_anchors:
                continue
            text = str(cell.value) if cell.value is not None else ""
            width = max(width, len(text))
        if column_letter:
            sheet.column_dimensions[column_letter].width = min(max(width + 2, min_width), max_width)


def write_workbook(path: Path, records: list[dict], rc_errors: list[tuple[str, str]]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    styles = {
        "bold": Font(bold=True),
        "note": Font(italic=True, color="9C0006"),
        "header_fill": PatternFill("solid", fgColor="DDEBF7"),
        "flag_fill": PatternFill("solid", fgColor="FFC7CE"),
        "center": Alignment(horizontal="center"),
        "link": Font(color="0563C1", underline="single"),
    }

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    used_titles = {"summary"}
    sheet_by_rc = {rc_path: workbook.create_sheet(_sheet_title(rc_path, used_titles)) for rc_path in RC_PATHS}

    records_by_rc: dict[str, list[dict]] = {}
    for record in records:
        records_by_rc.setdefault(record["rc_path"], []).append(record)

    dfm_anchor: dict[tuple[str, str], tuple[str, int]] = {}
    for rc_path in RC_PATHS:
        sheet = sheet_by_rc[rc_path]
        title_cell = sheet.cell(row=1, column=1, value=rc_path)
        title_cell.font = styles["bold"]
        row = 3
        for record in sorted(records_by_rc.get(rc_path, []), key=lambda r: r["name"].casefold()):
            header_row, next_row = _write_dfm_block(sheet, row, record, styles)
            dfm_anchor[(rc_path, record["name"])] = (sheet.title, header_row)
            row = next_row + 2
        _autosize(sheet, min_width=9, max_width=22)
        sheet.column_dimensions["A"].width = 26
        sheet.freeze_panes = "A3"

    summary_sheet.cell(
        row=1,
        column=1,
        value=(
            f'Project: {TARGET_PROJECT_NAME}    Ratio tolerance: {DECIMAL_PLACES} decimal places (±{TOLERANCE:g})'
            f'    Ultimate tolerance: {ULTIMATE_DECIMAL_PLACES} decimal places (±{ULTIMATE_TOLERANCE:g})'
            '    DFMs with "Adjusted" in their name are excluded'
        ),
    ).font = styles["bold"]
    header_row_index = 3
    headers = [
        "RC Path",
        "DFM Name",
        "Max Abs Ratio Diff",
        "Flagged Ratio Cells",
        "Max Abs Ultimate Diff",
        "Flagged Ultimate Cells",
        "Note",
    ]
    for col, text in enumerate(headers, start=1):
        cell = summary_sheet.cell(row=header_row_index, column=col, value=text)
        cell.font = styles["bold"]
    summary_sheet.freeze_panes = f"A{header_row_index + 1}"

    review_records = [record for record in records if record["needs_review"]]
    row = header_row_index + 1
    for record in sorted(review_records, key=lambda r: (r["rc_path"], r["name"].casefold())):
        summary_sheet.cell(row=row, column=1, value=record["rc_path"])
        name_cell = summary_sheet.cell(row=row, column=2, value=record["name"])
        anchor = dfm_anchor.get((record["rc_path"], record["name"]))
        if anchor:
            anchor_sheet_title, anchor_row = anchor
            name_cell.hyperlink = f"#'{anchor_sheet_title}'!A{anchor_row}"
            name_cell.font = styles["link"]
        if record["max_abs_diff"] is not None:
            diff_cell = summary_sheet.cell(row=row, column=3, value=record["max_abs_diff"])
            diff_cell.number_format = NUMBER_FORMAT
        summary_sheet.cell(row=row, column=4, value=record["flagged_cells"] or None)
        if record["ultimate_max_abs_diff"] is not None:
            ultimate_diff_cell = summary_sheet.cell(row=row, column=5, value=record["ultimate_max_abs_diff"])
            ultimate_diff_cell.number_format = ULTIMATE_NUMBER_FORMAT
        summary_sheet.cell(row=row, column=6, value=record["ultimate_flagged_cells"] or None)
        summary_sheet.cell(row=row, column=7, value=record["note"])
        row += 1

    for rc_path, note in rc_errors:
        summary_sheet.cell(row=row, column=1, value=rc_path)
        summary_sheet.cell(row=row, column=2, value="(reserving class)")
        summary_sheet.cell(row=row, column=7, value=note)
        row += 1

    if row == header_row_index + 1:
        summary_sheet.cell(row=row, column=1, value="No triangles need review.")

    _autosize(summary_sheet, min_width=12, max_width=80)

    descriptor, temporary_name = tempfile.mkstemp(prefix=f".{path.stem}-", suffix=".xlsx", dir=path.parent)
    os.close(descriptor)
    temporary_path = Path(temporary_name)
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, path)
    finally:
        workbook.close()
        if temporary_path.exists():
            temporary_path.unlink()


def main() -> int:
    records, rc_errors = run_comparison()
    write_workbook(OUTPUT_PATH, records, rc_errors)
    review_records = [record for record in records if record["needs_review"]]
    needs_attention = bool(review_records) or bool(rc_errors)
    print(f"Compared {len(RC_PATHS)} reserving classes, {len(records)} DFM(s) (Adjusted DFMs excluded).")
    print(f"{len(review_records)} DFM(s) need review" + (f", {len(rc_errors)} reserving class(es) could not be read" if rc_errors else "") + ".")
    print(f"Excel report: {OUTPUT_PATH}")
    if needs_attention:
        try:
            os.startfile(OUTPUT_PATH)  # noqa: S606 - opening the report just written, for the operator running this script
        except Exception as exc:
            print(f"Could not open the report automatically: {type(exc).__name__}: {exc}")
    return 0 if not needs_attention else 1


if __name__ == "__main__":
    raise SystemExit(main())
