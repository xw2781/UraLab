"""Combined ArcRho vs ResQ review for one reserving class: plain datasets + Result Selections.

Wraps ``dataset_side_by_side_review.py`` (plain triangles/vectors) and
``rs_dataset_side_by_side_review.py`` (Result Selection loaded datasets and
Selected Ultimate output) so a single target reserving class can be checked
in one run, with one combined workbook, instead of running both scripts and
opening two separate reports.

Nothing is written back to ArcRho or ResQ.

Run with Python 3.10 from the repository root, on a machine that can reach ResQ:

    py -3.10 python-api/migration/validation/combined_side_by_side_review.py --rc "Legacy\\HOL"
"""

from __future__ import annotations

import argparse
import os
import sys
import tempfile
from pathlib import Path

_VALIDATION_DIR = Path(__file__).resolve().parent
if str(_VALIDATION_DIR) not in sys.path:
    sys.path.insert(0, str(_VALIDATION_DIR))

import dataset_side_by_side_review as dsbs  # noqa: E402
import rs_dataset_side_by_side_review as rssbs  # noqa: E402

# Dataset/Result Selection names to leave out of the review entirely. A name is skipped
# if it CONTAINS any of these substrings (case-insensitive). Add more here as needed.
SKIP_DATASET_NAME_SUBSTRINGS: set[str] = {
    " - May 2026",
    "Growth Adjustment",
    "Accounting Cutoff",
}


def _is_skipped_name(name: str, skip_substrings: set[str]) -> bool:
    folded = name.casefold()
    return any(substring.casefold() in folded for substring in skip_substrings)


def _output_path(project_name: str, rc_paths: list[str]) -> Path:
    label = rc_paths[0].split("\\")[-1] if len(rc_paths) == 1 else f"{len(rc_paths)}rcs"
    safe_label = dsbs._INVALID_SHEET_CHARS.sub("-", label)
    return _VALIDATION_DIR / "results" / f"combined_side_by_side_{project_name}_{safe_label}.xlsx"


def _sheet_title(rc_path: str, suffix: str, used: set[str]) -> str:
    segments = [segment.strip() for segment in rc_path.split("\\") if segment.strip()]
    label = f"{segments[-3]} {segments[-1]}" if len(segments) >= 3 else rc_path
    label = dsbs._INVALID_SHEET_CHARS.sub("-", label)
    base = f"{label} {suffix}"[:31]
    title = base
    suffix_index = 2
    while title.casefold() in used:
        tail = f" ({suffix_index})"
        title = base[: 31 - len(tail)] + tail
        suffix_index += 1
    used.add(title.casefold())
    return title


def write_combined_workbook(
    path: Path,
    *,
    project_name: str,
    rc_paths: list[str],
    source_kinds: tuple[str, ...],
    dataset_records: list[dict],
    dataset_rc_errors: list[tuple[str, str]],
    rs_records: list[dict],
    rs_rc_errors: list[tuple[str, str]],
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    styles = {
        "bold": Font(bold=True),
        "note": Font(italic=True, color="9C0006"),
        "header_fill": PatternFill("solid", fgColor="DDEBF7"),
        "ultimate_fill": PatternFill("solid", fgColor="FFF2CC"),
        "flag_fill": PatternFill("solid", fgColor="FFC7CE"),
        "center": Alignment(horizontal="center"),
        "link": Font(color="0563C1", underline="single"),
    }

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    used_titles = {"summary"}

    dataset_records_by_rc: dict[str, list[dict]] = {}
    for record in dataset_records:
        dataset_records_by_rc.setdefault(record["rc_path"], []).append(record)
    rs_records_by_rc: dict[str, list[dict]] = {}
    for record in rs_records:
        rs_records_by_rc.setdefault(record["rc_path"], []).append(record)

    dataset_anchors: dict[tuple[str, str, str], tuple[str, int]] = {}
    rs_anchors: dict[tuple[str, str], tuple[str, int]] = {}

    for rc_path in rc_paths:
        ds_records = sorted(
            dataset_records_by_rc.get(rc_path, []), key=lambda r: (r["kind"], r["name"].casefold())
        )
        if ds_records:
            sheet = workbook.create_sheet(_sheet_title(rc_path, "DS", used_titles))
            sheet.cell(row=1, column=1, value=rc_path).font = styles["bold"]
            row = 3
            for record in ds_records:
                header_row, next_row = dsbs._write_dataset_block(sheet, row, record, styles)
                dataset_anchors[(rc_path, record["kind"], record["name"])] = (sheet.title, header_row)
                row = next_row + 2
            dsbs._autosize(sheet, min_width=9, max_width=22)
            sheet.column_dimensions["A"].width = 30
            sheet.freeze_panes = "A3"

        rs_recs = sorted(rs_records_by_rc.get(rc_path, []), key=lambda r: r["name"].casefold())
        if rs_recs:
            sheet = workbook.create_sheet(_sheet_title(rc_path, "RS", used_titles))
            sheet.cell(row=1, column=1, value=rc_path).font = styles["bold"]
            row = 3
            for record in rs_recs:
                header_row, next_row = rssbs._write_rs_block(sheet, row, record, styles)
                rs_anchors[(rc_path, record["name"])] = (sheet.title, header_row)
                row = next_row + 2
            rssbs._autosize(sheet, min_width=9, max_width=22)
            sheet.column_dimensions["A"].width = 26
            sheet.freeze_panes = "A3"

    summary_sheet.cell(
        row=1,
        column=1,
        value=(
            f"Project: {project_name}    Reserving class(es): {', '.join(rc_paths)}    "
            f"Plain-dataset scope: {', '.join(source_kinds)}"
        ),
    ).font = styles["bold"]
    headers = ["Type", "RC Path", "Kind", "Name", "Max Abs Diff", "Flagged Cells", "Note"]
    for col, text in enumerate(headers, start=1):
        summary_sheet.cell(row=3, column=col, value=text).font = styles["bold"]
    summary_sheet.freeze_panes = "A4"

    row = 4
    for record in sorted(
        (r for r in dataset_records if r["needs_review"]),
        key=lambda r: (r["rc_path"], r["kind"], r["name"].casefold()),
    ):
        summary_sheet.cell(row=row, column=1, value="Dataset")
        summary_sheet.cell(row=row, column=2, value=record["rc_path"])
        summary_sheet.cell(row=row, column=3, value=record["kind"])
        name_cell = summary_sheet.cell(row=row, column=4, value=record["name"])
        anchor = dataset_anchors.get((record["rc_path"], record["kind"], record["name"]))
        if anchor:
            anchor_sheet_title, anchor_row = anchor
            name_cell.hyperlink = f"#'{anchor_sheet_title}'!A{anchor_row}"
            name_cell.font = styles["link"]
        if record["max_abs_diff"] is not None:
            summary_sheet.cell(row=row, column=5, value=record["max_abs_diff"]).number_format = dsbs.NUMBER_FORMAT
        summary_sheet.cell(row=row, column=6, value=record["flagged_cells"] or None)
        summary_sheet.cell(row=row, column=7, value=record["note"])
        row += 1

    for record in sorted(
        (r for r in rs_records if r["needs_review"]), key=lambda r: (r["rc_path"], r["name"].casefold())
    ):
        summary_sheet.cell(row=row, column=1, value="Result Selection")
        summary_sheet.cell(row=row, column=2, value=record["rc_path"])
        name_cell = summary_sheet.cell(row=row, column=4, value=record["name"])
        anchor = rs_anchors.get((record["rc_path"], record["name"]))
        if anchor:
            anchor_sheet_title, anchor_row = anchor
            name_cell.hyperlink = f"#'{anchor_sheet_title}'!A{anchor_row}"
            name_cell.font = styles["link"]
        if record["max_abs_diff"] is not None:
            summary_sheet.cell(row=row, column=5, value=record["max_abs_diff"]).number_format = rssbs.NUMBER_FORMAT
        summary_sheet.cell(row=row, column=6, value=record["flagged_cells"] or None)
        summary_sheet.cell(row=row, column=7, value=record["note"])
        row += 1

    for rc_path, note in dataset_rc_errors:
        summary_sheet.cell(row=row, column=1, value="Dataset")
        summary_sheet.cell(row=row, column=2, value=rc_path)
        summary_sheet.cell(row=row, column=3, value="(reserving class)")
        summary_sheet.cell(row=row, column=7, value=note)
        row += 1
    for rc_path, note in rs_rc_errors:
        summary_sheet.cell(row=row, column=1, value="Result Selection")
        summary_sheet.cell(row=row, column=2, value=rc_path)
        summary_sheet.cell(row=row, column=3, value="(reserving class)")
        summary_sheet.cell(row=row, column=7, value=note)
        row += 1

    if row == 4:
        summary_sheet.cell(row=row, column=1, value="Nothing needs review.")

    dsbs._autosize(summary_sheet, min_width=12, max_width=80)

    descriptor, temporary_name = tempfile.mkstemp(prefix=f".{path.stem}-", suffix=".xlsx", dir=path.parent)
    os.close(descriptor)
    temporary_path = Path(temporary_name)
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, path)
    finally:
        workbook.close()
        if temporary_path.exists():
            temporary_path.unlink()


def _parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--project", default=dsbs.TARGET_PROJECT_NAME, help="ResQ project name to review.")
    parser.add_argument(
        "--rc",
        action="append",
        required=True,
        help="Only review reserving classes whose path contains this text; repeatable.",
    )
    parser.add_argument(
        "--source-kind",
        action="append",
        choices=[*dsbs.SOURCE_KINDS, "all"],
        help="Plain-dataset source kind to include; repeatable. Defaults to all kinds for a targeted RC check.",
    )
    parser.add_argument("--no-open", action="store_true", help="Do not open the workbook when the run finishes.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = _parse_args(argv)

    needles = [text.casefold() for text in args.rc]
    rc_paths = [path for path in dsbs.RC_PATHS if any(needle in path.casefold() for needle in needles)]
    if not rc_paths:
        print("No reserving class matched --rc.")
        return 2

    selected = args.source_kind or ["all"]
    source_kinds = dsbs.SOURCE_KINDS if "all" in selected else tuple(dict.fromkeys(selected))

    print(f"Reserving class(es): {', '.join(rc_paths)}")
    print("Comparing plain datasets...")
    dataset_records, dataset_rc_errors = dsbs.run_comparison(
        project_name=args.project, rc_paths=rc_paths, source_kinds=source_kinds
    )

    print("Comparing Result Selections...")
    rs_records, rs_rc_errors = rssbs.run_comparison(project_name=args.project, rc_paths=rc_paths)

    if SKIP_DATASET_NAME_SUBSTRINGS:
        skipped_dataset_count = sum(
            1 for r in dataset_records if _is_skipped_name(r["name"], SKIP_DATASET_NAME_SUBSTRINGS)
        )
        skipped_rs_count = sum(1 for r in rs_records if _is_skipped_name(r["name"], SKIP_DATASET_NAME_SUBSTRINGS))
        dataset_records = [
            r for r in dataset_records if not _is_skipped_name(r["name"], SKIP_DATASET_NAME_SUBSTRINGS)
        ]
        rs_records = [r for r in rs_records if not _is_skipped_name(r["name"], SKIP_DATASET_NAME_SUBSTRINGS)]
        print(f"Skipping {skipped_dataset_count} dataset(s) and {skipped_rs_count} Result Selection(s) by name.")

    output_path = _output_path(args.project, rc_paths)
    write_combined_workbook(
        output_path,
        project_name=args.project,
        rc_paths=rc_paths,
        source_kinds=source_kinds,
        dataset_records=dataset_records,
        dataset_rc_errors=dataset_rc_errors,
        rs_records=rs_records,
        rs_rc_errors=rs_rc_errors,
    )

    dataset_needs_review = [r for r in dataset_records if r["needs_review"]]
    rs_needs_review = [r for r in rs_records if r["needs_review"]]
    needs_attention = (
        bool(dataset_needs_review) or bool(rs_needs_review) or bool(dataset_rc_errors) or bool(rs_rc_errors)
    )

    print(f"Plain datasets: {len(dataset_records)} compared, {len(dataset_needs_review)} need review.")
    print(f"Result Selections: {len(rs_records)} compared, {len(rs_needs_review)} need review.")
    print(f"Excel report: {output_path}")
    if needs_attention and not args.no_open:
        try:
            os.startfile(output_path)  # noqa: S606 - opening the report just written, for the operator running this script
        except Exception as exc:
            print(f"Could not open the report automatically: {type(exc).__name__}: {exc}")
    return 0 if not needs_attention else 1


if __name__ == "__main__":
    raise SystemExit(main())
