"""Side-by-side ArcRho vs ResQ plain-dataset review workbook.

For every plain triangle and vector -- a dataset with no method behind it --
in the 17 reserving-class paths of NJ_Annual_Prod_2026 Q3-Aug, this lays the
persisted ArcRho values and the live ResQ values next to each other, with a
difference matrix (ArcRho minus ResQ) to their right: one block per dataset,
grouped onto one sheet per reserving class.

Scope defaults to the datasets ArcRho's own data engine rebuilds from source
data (``source_kind == "engine"``), which is where the two systems can
genuinely disagree. ``--source-kind`` widens it to the calculated and the
plain input datasets as well. Method outputs -- DFM, BF, Cape Cod, Result
Selection, Berquist-Sherman, Bootstrap -- are never included; the DFM ratio
and Result Selection reviews beside this file cover those.

A Summary sheet links straight to every dataset that disagrees by more than
tolerance, whose shape differs, or that is missing from one side, and a
Coverage sheet lists every dataset compared so the reviewer can confirm what
the run actually looked at.

Nothing is written back to ArcRho or ResQ. The ArcRho side is read from the
persisted CSV cache exactly as the app serves it; ``validate_engine_resq_parity.py``
beside this file is the tool that instead regenerates each dataset from the
engine first.

Run with Python 3.10 from the repository root, on a machine that can reach ResQ:

    py -3.10 python-api/migration/validation/dataset_side_by_side_review.py
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import tempfile
from pathlib import Path
from typing import Any

_VALIDATION_DIR = Path(__file__).resolve().parent
_MIGRATION_DIR = _VALIDATION_DIR.parent
if str(_MIGRATION_DIR) not in sys.path:
    sys.path.insert(0, str(_MIGRATION_DIR))

import resq_data_migration as migration  # noqa: E402
from arcrho_api.sidecar_core_contract import display_lengths, stored_lengths  # noqa: E402
from arcrho_api.triangle_rollup import rollup_reason, rollup_triangle  # noqa: E402
from resq_migration.catalog import _triangle_source_kind  # noqa: E402
from resq_migration.core import (  # noqa: E402
    METHOD_TYPE_NONE_CODE,
    _clean_name,
    _encode_rc_folder,
    _normalize_import_name,
    _safe_attr,
)
from resq_migration.engine_parity import read_engine_csv  # noqa: E402
from resq_migration.extractors import export_triangle, export_vector  # noqa: E402


TARGET_PROJECT_NAME = "NJ_Annual_Prod_2026 Q3-Aug"
RC_PATHS = [
    r"PRNJ - PA\PA\NY\Direct Group\BI Total",
    r"PRNJ - PA\PA\NY\Direct Group\MP+PIP",
    r"PRNJ - PA\PA\Penn+CT\Direct Group\BI Total",
    r"PRNJ - PA\PA\Penn+CT\Direct Group\MP+PIP",
    r"PRNJ - PA\PA\All States\Direct Group\PD+UMPD",
    r"PRNJ - PA\PA\All States\Direct Group\COL",
    r"PRNJ - PA\PA\All States\Direct Group\CMPxCAT",
    r"PRNJ - PA\PA\NJ\Direct Group\MP+PIP",
    r"PRNJ - PA\PA\NJ\Direct Group\BIR51+UMBIR51",
    r"PRNJ - PA\PA\NJ\Direct Group\BIx51+UMBIx51",
    r"HPPREF\HO+DF\NJ\Legacy\HOL",
    r"HPPREF\HO+DF\NJ\Legacy\HOPxCAT",
    r"Rider\MC\All States\Direct Group\BI+PIP",
    r"Rider\MC\All States\Direct Group\PD+UMPD",
    r"Rider\MC\All States\Direct Group\PhysDxCat",
    r"PRNJ - PA\PA\MA\Direct Group\BI Total",
    r"PRNJ - PA\PA\MA\Direct Group\MP+PIP",
]

SOURCE_KINDS = ("engine", "calculated", "input")
DEFAULT_SOURCE_KINDS = ("engine",)

TRIANGLE_KIND = "Triangle"
VECTOR_KIND = "Vector"
VECTOR_COLUMN_LABEL = "Value"

# Plain datasets span wildly different magnitudes on one sheet -- claim counts
# in the tens next to earned premium in the millions and development ratios
# near 1 -- so a single absolute tolerance either misses real mismatches on
# small vectors or drowns in float noise on large ones. The absolute floor is
# the ResQ import's own rule, that two cells agree when they round to the same
# two decimals; the relative floor keeps ordinary double-precision noise on
# eight-figure values from being reported as a difference.
ABS_TOLERANCE = 0.005
REL_TOLERANCE = 1e-9
NUMBER_FORMAT = "#,##0.0000"

_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


def _output_path(project_name: str) -> Path:
    return _VALIDATION_DIR / "results" / f"dataset_side_by_side_{project_name}.xlsx"


def _tolerance(a_val: float, r_val: float) -> float:
    scale = max(abs(a_val), abs(r_val))
    return max(ABS_TOLERANCE, REL_TOLERANCE * scale)


def _ym_to_month_index(raw: Any) -> int | None:
    digits = "".join(ch for ch in str(raw or "") if ch.isdigit())
    if len(digits) < 6:
        return None
    year, month = int(digits[:4]), int(digits[4:6])
    if year <= 0 or month < 1 or month > 12:
        return None
    return year * 12 + (month - 1)


def _valuation_months(project_name: str) -> int | None:
    """Months from the project's Origin Start Date through its Development End Date.

    The same anchor ``dataset_service.valuation_months`` uses to roll a
    hand-entered dataset's stored rows up to its display shape -- read here
    straight from ``general_settings.json`` so this script does not have to
    import the app-server's FastAPI-dependent service module.
    """

    path = migration.PROJECT_DATA_DIR.parent / "general_settings.json"
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None
    if not isinstance(payload, dict):
        return None
    origin_start = _ym_to_month_index(payload.get("origin_start_date"))
    development_end = _ym_to_month_index(payload.get("development_end_date"))
    if origin_start is None or development_end is None:
        return None
    return development_end - origin_start + 1


# --------------------------------------------------------------------------
# ArcRho side: the persisted sidecars and their CSV cache
# --------------------------------------------------------------------------


def _read_arcrho_datasets(
    rc_dir: Path, source_kinds: tuple[str, ...], valuation_months: int | None
) -> dict[tuple[str, str], dict]:
    """Map (kind, name) -> persisted ArcRho dataset for one reserving class.

    Only sidecars with no method behind them are returned, filtered to the
    requested source kinds. The values are read from the sidecar's own
    ``csv_file``, which is written at the dataset's *stored* shape -- ResQ
    always shows a dataset at its *display* shape, so a hand-entered dataset
    stored coarser than it is displayed is rolled up here the same way
    ``dataset_service._display_view_of_stored_values`` does it for the app's
    own Dataset window, or the two sides would be compared at different
    granularities. An engine or calculated dataset is never rolled up here,
    matching the app: its own cache is regenerated at the shape it is asked
    for, and its sidecar's "stored" fields name the source-table granularity
    rather than that cache's own shape, so rolling it up from those fields
    would aggregate against the wrong source shape.
    """

    sidecar_dir = rc_dir / "sidecars"
    dataset_dir = rc_dir / "datasets"
    out: dict[tuple[str, str], dict] = {}
    if not sidecar_dir.is_dir():
        return out
    for path in sorted(sidecar_dir.glob("*.json")):
        try:
            meta = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        if not isinstance(meta, dict):
            continue
        if _clean_name(meta.get("method_type") or "None").lower() not in {"", "none"}:
            continue
        source_kind = _clean_name(meta.get("source_kind")) or "input"
        if source_kind not in source_kinds:
            continue
        name = _normalize_import_name(meta.get("dataset_name")) or path.stem
        kind = VECTOR_KIND if _clean_name(meta.get("data_format")).lower() == "vector" else TRIANGLE_KIND

        values: list[list[float | None]] | None = None
        error = ""
        csv_name = _clean_name(meta.get("csv_file"))
        csv_path = dataset_dir / csv_name if csv_name else None
        stored_origin, stored_development = stored_lengths(meta)
        display_origin, display_development = display_lengths(meta)
        if csv_path is None:
            error = "sidecar names no CSV file"
        elif not csv_path.is_file():
            error = f"cached CSV {csv_name!r} is missing"
        else:
            try:
                values = read_engine_csv(csv_path)
            except Exception as exc:
                error = f"could not read {csv_name!r}: {type(exc).__name__}: {exc}"
            if (
                values is not None
                and source_kind == "input"
                and (display_origin, display_development) not in {(0, 0), (stored_origin, stored_development)}
            ):
                # A vector holds one column, so only its rows are aggregated: a
                # plain block sum, the calendar form of the roll-up. This mirrors
                # precedent_cache_service._rollup_arguments's vector branch --
                # rolling a vector up by its (period, period) display shape like
                # a triangle would ask for a square block instead of one column.
                if kind == VECTOR_KIND:
                    rollup_target_development = stored_development
                    cumulative = True
                    calendar = True
                else:
                    rollup_target_development = display_development
                    cumulative = bool(meta.get("cumulative", True))
                    calendar = bool(meta.get("calendar", False))
                reason = rollup_reason(
                    stored_origin, stored_development, display_origin, rollup_target_development, calendar=calendar
                )
                if reason:
                    error = f"cannot roll up to display shape: {reason}"
                elif not valuation_months:
                    error = "cannot roll up to display shape: project valuation date is unavailable"
                else:
                    try:
                        values = rollup_triangle(
                            values,
                            source_origin_length=stored_origin,
                            source_development_length=stored_development,
                            target_origin_length=display_origin,
                            target_development_length=rollup_target_development,
                            valuation_months=valuation_months,
                            cumulative=cumulative,
                            calendar=calendar,
                        )
                    except Exception as exc:
                        error = f"could not roll up to display shape: {type(exc).__name__}: {exc}"

        out[(kind, name)] = {
            "name": name,
            "kind": kind,
            "source_kind": source_kind,
            "dataset_type": _normalize_import_name(meta.get("dataset_type")) or name,
            "origin_length": display_origin or stored_origin,
            "development_length": display_development or stored_development,
            "values": values,
            "error": error,
        }
    return out


# --------------------------------------------------------------------------
# ResQ side: the live COM collections
# --------------------------------------------------------------------------


def _resq_method_type_code(dataset: Any) -> int:
    try:
        return int(_safe_attr(dataset, "MethodType", METHOD_TYPE_NONE_CODE))
    except (TypeError, ValueError):
        return METHOD_TYPE_NONE_CODE


def _resq_dataset_type_name(dataset: Any) -> str:
    return _normalize_import_name(_safe_attr(_safe_attr(dataset, "DatasetType", None), "Name", ""))


def _read_resq_datasets(
    reserving_class: Any,
    source_kinds: tuple[str, ...],
    progress,
) -> tuple[dict[tuple[str, str], dict], list[tuple[str, str, str]]]:
    """Map (kind, name) -> exported ResQ payload, plus any per-dataset read errors."""

    out: dict[tuple[str, str], dict] = {}
    errors: list[tuple[str, str, str]] = []
    for kind, collection_name, exporter in (
        (TRIANGLE_KIND, "Triangles", export_triangle),
        (VECTOR_KIND, "Vectors", export_vector),
    ):
        try:
            collection = list(getattr(reserving_class, collection_name)())
        except Exception as exc:
            errors.append((kind, "", f"could not list ResQ {collection_name}: {type(exc).__name__}: {exc}"))
            continue
        for dataset in collection:
            name = _normalize_import_name(_safe_attr(dataset, "Name", ""))
            if not name:
                continue
            # A dataset carrying a method type is that method's output, not a
            # plain dataset, so it is out of scope whatever its name suggests.
            if _resq_method_type_code(dataset) != METHOD_TYPE_NONE_CODE:
                continue
            dataset_type = _resq_dataset_type_name(dataset) or name
            if _triangle_source_kind(name, dataset_type) not in source_kinds:
                continue
            try:
                payload = exporter(dataset)
            except Exception as exc:
                errors.append((kind, name, f"could not read from ResQ: {type(exc).__name__}: {exc}"))
                continue
            progress(f"    {kind.lower()}: {name}")
            out[(kind, name)] = payload
    return out, errors


# --------------------------------------------------------------------------
# Comparison
# --------------------------------------------------------------------------


def _matrix_shape(matrix: list[list[Any]] | None) -> tuple[int, int]:
    if not matrix:
        return (0, 0)
    return (len(matrix), max((len(row) for row in matrix), default=0))


def _cell(matrix: list[list[Any]] | None, row: int, column: int) -> float | None:
    if not matrix or row >= len(matrix) or column >= len(matrix[row]):
        return None
    value = matrix[row][column]
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _is_ordinal_labels(labels: list[str]) -> bool:
    """True when the labels are nothing but the 1..n row numbers ResQ falls back to."""

    return not labels or labels == [str(i + 1) for i in range(len(labels))]


def _origin_labels_by_row_count(resq_datasets: dict[tuple[str, str], dict]) -> dict[int, list[str]]:
    """Origin labels a triangle in this reserving class carries, keyed by its row count.

    ResQ exposes no origin label on a Vector, so a vector block would otherwise
    be numbered 1..n. A triangle of the same height in the same reserving class
    spans the same origin periods, so its labels name the vector's rows too.
    """

    out: dict[int, list[str]] = {}
    for (kind, _name), payload in resq_datasets.items():
        if kind != TRIANGLE_KIND:
            continue
        labels = [str(label) for label in payload.get("origin_labels", []) or []]
        if len(labels) not in out and not _is_ordinal_labels(labels):
            out[len(labels)] = labels
    return out


def _build_record(
    rc_path: str,
    kind: str,
    name: str,
    arcrho: dict | None,
    resq_payload: dict | None,
    origin_label_fallback: dict[int, list[str]] | None = None,
) -> dict:
    note_parts: list[str] = []

    arcrho_values = arcrho.get("values") if arcrho else None
    if arcrho is not None and arcrho.get("error"):
        note_parts.append(arcrho["error"])
    if arcrho is None:
        note_parts.append(f"{kind.lower()} exists in ResQ but ArcRho holds no sidecar for it")
    if resq_payload is None:
        note_parts.append(f"{kind.lower()} is persisted in ArcRho but was not found in ResQ")

    resq_values = resq_payload.get("values") if resq_payload else None
    origin_labels = [str(label) for label in (resq_payload or {}).get("origin_labels", []) or []]
    if kind == VECTOR_KIND:
        dev_labels = [VECTOR_COLUMN_LABEL]
    else:
        dev_labels = [str(label) for label in (resq_payload or {}).get("development_labels", []) or []]

    arcrho_shape = _matrix_shape(arcrho_values)
    resq_shape = _matrix_shape(resq_values)
    if arcrho_values is not None and resq_values is not None and arcrho_shape != resq_shape:
        note_parts.append(
            f"shape differs: ArcRho {arcrho_shape[0]}x{arcrho_shape[1]}, ResQ {resq_shape[0]}x{resq_shape[1]}"
        )

    row_count = max(arcrho_shape[0], resq_shape[0])
    column_count = max(arcrho_shape[1], resq_shape[1], 1)
    if row_count and _is_ordinal_labels(origin_labels):
        borrowed = (origin_label_fallback or {}).get(row_count)
        if borrowed:
            origin_labels = list(borrowed)
    for i in range(len(origin_labels), row_count):
        origin_labels.append(str(i + 1))
    for j in range(len(dev_labels), column_count):
        dev_labels.append(str(j + 1))

    arcrho_matrix: list[list[float | None]] = []
    resq_matrix: list[list[float | None]] = []
    diff_matrix: list[list[float | None]] = []
    max_abs_diff: float | None = None
    max_rel_diff: float | None = None
    flagged_cells = 0
    only_arcrho_cells = 0
    only_resq_cells = 0

    both_present = arcrho_values is not None and resq_values is not None
    for i in range(row_count):
        a_row: list[float | None] = []
        r_row: list[float | None] = []
        d_row: list[float | None] = []
        for j in range(column_count):
            a_val = _cell(arcrho_values, i, j)
            r_val = _cell(resq_values, i, j)
            a_row.append(a_val)
            r_row.append(r_val)
            if a_val is None or r_val is None:
                d_row.append(None)
                if both_present:
                    if a_val is not None:
                        only_arcrho_cells += 1
                    elif r_val is not None:
                        only_resq_cells += 1
                continue
            diff = a_val - r_val
            d_row.append(diff)
            scale = max(abs(a_val), abs(r_val))
            relative = 0.0 if scale == 0 else abs(diff) / scale
            if max_abs_diff is None or abs(diff) > max_abs_diff:
                max_abs_diff = abs(diff)
            if max_rel_diff is None or relative > max_rel_diff:
                max_rel_diff = relative
            if abs(diff) > _tolerance(a_val, r_val):
                flagged_cells += 1
        arcrho_matrix.append(a_row)
        resq_matrix.append(r_row)
        diff_matrix.append(d_row)

    if only_arcrho_cells:
        note_parts.append(f"{only_arcrho_cells} cell(s) with a value in ArcRho only")
    if only_resq_cells:
        note_parts.append(f"{only_resq_cells} cell(s) with a value in ResQ only")

    note = "; ".join(part for part in note_parts if part)
    needs_review = bool(note) or flagged_cells > 0

    source_kind = (arcrho or {}).get("source_kind") or (
        _triangle_source_kind(name, (resq_payload or {}).get("dataset_type") or name)
    )

    return {
        "rc_path": rc_path,
        "kind": kind,
        "name": name,
        "source_kind": source_kind,
        "origin_labels": origin_labels,
        "dev_labels": dev_labels,
        "arcrho_matrix": arcrho_matrix,
        "resq_matrix": resq_matrix,
        "diff_matrix": diff_matrix,
        "row_count": row_count,
        "column_count": column_count,
        "arcrho_shape": arcrho_shape,
        "resq_shape": resq_shape,
        "max_abs_diff": max_abs_diff,
        "max_rel_diff": max_rel_diff,
        "flagged_cells": flagged_cells,
        "note": note,
        "needs_review": needs_review,
    }


def run_comparison(
    *,
    project_name: str = TARGET_PROJECT_NAME,
    rc_paths: list[str] | None = None,
    source_kinds: tuple[str, ...] = DEFAULT_SOURCE_KINDS,
    app_factory=None,
    progress=print,
) -> tuple[list[dict], list[tuple[str, str]]]:
    """Compare every in-scope plain dataset.

    Returns (records, rc_errors) where records covers every dataset found on
    either side and rc_errors lists reserving classes ResQ itself refused.
    """

    if app_factory is None:
        try:
            import win32com.client
        except ImportError as exc:
            raise RuntimeError("pywin32 is required: pip install pywin32") from exc

    rc_paths = list(rc_paths if rc_paths is not None else RC_PATHS)
    previous_scope = migration._apply_runtime_scope(project_name, migration.SERVER_ROOT)
    valuation_months = _valuation_months(project_name)
    app = app_factory() if app_factory is not None else win32com.client.Dispatch("ResQ3Automation.ResQApplication")
    records: list[dict] = []
    rc_errors: list[tuple[str, str]] = []
    try:
        app.ConnectByName(migration.CONNECTION_NAME, migration.USER_NAME, migration.PASSWORD)
        project = app.Projects().Item(project_name)

        for rc_index, rc_path in enumerate(rc_paths, start=1):
            progress(f"RC {rc_index}/{len(rc_paths)}: {rc_path}")
            rc_dir = migration.PROJECT_DATA_DIR / _encode_rc_folder(rc_path)
            arcrho_datasets = _read_arcrho_datasets(rc_dir, source_kinds, valuation_months)

            try:
                reserving_class = project.ReservingClasses().Item(rc_path)
            except Exception as exc:
                rc_errors.append((rc_path, f"could not read ResQ reserving class: {type(exc).__name__}: {exc}"))
                continue

            resq_datasets, read_errors = _read_resq_datasets(reserving_class, source_kinds, progress)
            for kind, name, message in read_errors:
                rc_errors.append((rc_path, f"{kind} {name or '(collection)'}: {message}"))

            origin_label_fallback = _origin_labels_by_row_count(resq_datasets)
            keys = sorted(
                set(arcrho_datasets) | set(resq_datasets),
                key=lambda key: (key[0], key[1].casefold()),
            )
            for kind, name in keys:
                records.append(
                    _build_record(
                        rc_path,
                        kind,
                        name,
                        arcrho_datasets.get((kind, name)),
                        resq_datasets.get((kind, name)),
                        origin_label_fallback,
                    )
                )
    finally:
        try:
            app.Disconnect()
        except Exception:
            pass
        migration._restore_runtime_scope(previous_scope)
    return records, rc_errors


# --------------------------------------------------------------------------
# Workbook
# --------------------------------------------------------------------------


def _sheet_title(rc_path: str, used: set[str]) -> str:
    segments = [segment.strip() for segment in rc_path.split("\\") if segment.strip()]
    label = f"{segments[-3]} {segments[-1]}" if len(segments) >= 3 else rc_path
    label = _INVALID_SHEET_CHARS.sub("-", label)[:31]
    base = label
    suffix_index = 2
    while label.casefold() in used:
        suffix = f" ({suffix_index})"
        label = base[: 31 - len(suffix)] + suffix
        suffix_index += 1
    used.add(label.casefold())
    return label


def _write_dataset_block(sheet, start_row: int, record: dict, styles: dict) -> tuple[int, int]:
    """Write one dataset's ArcRho | ResQ | Diff matrices. Returns (header_row, next_free_row)."""

    column_count = max(record["column_count"], 1)
    row_count = record["row_count"]
    gap = 1

    header_row = start_row
    header_cell = sheet.cell(row=header_row, column=1, value=record["name"])
    header_cell.font = styles["bold"]
    if record["note"]:
        sheet.cell(row=header_row, column=2, value=record["note"]).font = styles["note"]

    group_row = header_row + 1
    label_row = header_row + 2
    data_start_row = header_row + 3

    origin_col = 1
    arcrho_start_col = 2
    resq_start_col = arcrho_start_col + column_count + gap
    diff_start_col = resq_start_col + column_count + gap

    def _group_header(col: int, text: str) -> None:
        cell = sheet.cell(row=group_row, column=col, value=text)
        cell.font = styles["bold"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center"]
        if column_count > 1:
            sheet.merge_cells(
                start_row=group_row, start_column=col, end_row=group_row, end_column=col + column_count - 1
            )

    _group_header(arcrho_start_col, "ArcRho")
    _group_header(resq_start_col, "ResQ")
    _group_header(diff_start_col, "Diff (ArcRho − ResQ)")

    sheet.cell(row=label_row, column=origin_col, value="Origin").font = styles["bold"]
    for j in range(column_count):
        label = record["dev_labels"][j] if j < len(record["dev_labels"]) else str(j + 1)
        for base_col in (arcrho_start_col, resq_start_col, diff_start_col):
            sheet.cell(row=label_row, column=base_col + j, value=label).font = styles["bold"]

    for i in range(row_count):
        row = data_start_row + i
        origin_label = record["origin_labels"][i] if i < len(record["origin_labels"]) else i + 1
        sheet.cell(row=row, column=origin_col, value=origin_label)
        a_row = record["arcrho_matrix"][i] if i < len(record["arcrho_matrix"]) else []
        r_row = record["resq_matrix"][i] if i < len(record["resq_matrix"]) else []
        d_row = record["diff_matrix"][i] if i < len(record["diff_matrix"]) else []
        for j in range(column_count):
            a_val = a_row[j] if j < len(a_row) else None
            r_val = r_row[j] if j < len(r_row) else None
            d_val = d_row[j] if j < len(d_row) else None
            flagged = (
                d_val is not None
                and a_val is not None
                and r_val is not None
                and abs(d_val) > _tolerance(a_val, r_val)
            )
            a_cell = sheet.cell(row=row, column=arcrho_start_col + j, value=a_val)
            r_cell = sheet.cell(row=row, column=resq_start_col + j, value=r_val)
            d_cell = sheet.cell(row=row, column=diff_start_col + j, value=d_val)
            for cell in (a_cell, r_cell, d_cell):
                cell.number_format = NUMBER_FORMAT
                if flagged:
                    cell.fill = styles["flag_fill"]

    return header_row, data_start_row + max(row_count, 1)


def _autosize(sheet, *, min_width: int, max_width: int) -> None:
    # A merged group header ("ArcRho" / "ResQ" / "Diff (ArcRho − ResQ)") is
    # anchored on one cell but reads across every column of that dataset's
    # block. Different blocks on the same sheet merge different spans, so
    # over many blocks nearly every column ends up anchoring some block's
    # header at some point; counting that text toward a single column's width
    # would balloon every data column to fit "Diff (ArcRho − ResQ)".
    wide_merge_anchors = {
        (merged_range.min_row, merged_range.min_col)
        for merged_range in sheet.merged_cells.ranges
        if merged_range.max_col > merged_range.min_col
    }
    for column_cells in sheet.columns:
        column_letter = None
        width = 0
        for cell in column_cells:
            letter = getattr(cell, "column_letter", None)
            if letter is None:
                continue
            column_letter = letter
            if (cell.row, cell.column) in wide_merge_anchors:
                continue
            text = str(cell.value) if cell.value is not None else ""
            width = max(width, len(text))
        if column_letter:
            sheet.column_dimensions[column_letter].width = min(max(width + 2, min_width), max_width)


def _shape_text(shape: tuple[int, int]) -> str:
    return "" if shape == (0, 0) else f"{shape[0]}x{shape[1]}"


def write_workbook(
    path: Path,
    records: list[dict],
    rc_errors: list[tuple[str, str]],
    *,
    project_name: str,
    rc_paths: list[str],
    source_kinds: tuple[str, ...],
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    styles = {
        "bold": Font(bold=True),
        "note": Font(italic=True, color="9C0006"),
        "header_fill": PatternFill("solid", fgColor="DDEBF7"),
        "flag_fill": PatternFill("solid", fgColor="FFC7CE"),
        "center": Alignment(horizontal="center"),
        "link": Font(color="0563C1", underline="single"),
    }

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    coverage_sheet = workbook.create_sheet("Coverage")

    used_titles = {"summary", "coverage"}
    sheet_by_rc = {rc_path: workbook.create_sheet(_sheet_title(rc_path, used_titles)) for rc_path in rc_paths}

    records_by_rc: dict[str, list[dict]] = {}
    for record in records:
        records_by_rc.setdefault(record["rc_path"], []).append(record)

    anchors: dict[tuple[str, str, str], tuple[str, int]] = {}
    for rc_path in rc_paths:
        sheet = sheet_by_rc[rc_path]
        sheet.cell(row=1, column=1, value=rc_path).font = styles["bold"]
        row = 3
        for record in sorted(records_by_rc.get(rc_path, []), key=lambda r: (r["kind"], r["name"].casefold())):
            header_row, next_row = _write_dataset_block(sheet, row, record, styles)
            anchors[(rc_path, record["kind"], record["name"])] = (sheet.title, header_row)
            row = next_row + 2
        _autosize(sheet, min_width=9, max_width=22)
        sheet.column_dimensions["A"].width = 30
        sheet.freeze_panes = "A3"

    summary_sheet.cell(
        row=1,
        column=1,
        value=(
            f"Project: {project_name}    Scope: plain {', '.join(source_kinds)} triangles and vectors    "
            f"Tolerance: max({ABS_TOLERANCE:g}, {REL_TOLERANCE:g} * |value|)"
        ),
    ).font = styles["bold"]
    headers = [
        "RC Path",
        "Kind",
        "Dataset",
        "Source",
        "ArcRho Shape",
        "ResQ Shape",
        "Max Abs Diff",
        "Max Rel Diff",
        "Flagged Cells",
        "Note",
    ]
    for col, text in enumerate(headers, start=1):
        summary_sheet.cell(row=3, column=col, value=text).font = styles["bold"]
    summary_sheet.freeze_panes = "A4"

    review_records = [record for record in records if record["needs_review"]]
    row = 4
    for record in sorted(review_records, key=lambda r: (r["rc_path"], r["kind"], r["name"].casefold())):
        summary_sheet.cell(row=row, column=1, value=record["rc_path"])
        summary_sheet.cell(row=row, column=2, value=record["kind"])
        name_cell = summary_sheet.cell(row=row, column=3, value=record["name"])
        anchor = anchors.get((record["rc_path"], record["kind"], record["name"]))
        if anchor:
            anchor_sheet_title, anchor_row = anchor
            name_cell.hyperlink = f"#'{anchor_sheet_title}'!A{anchor_row}"
            name_cell.font = styles["link"]
        summary_sheet.cell(row=row, column=4, value=record["source_kind"])
        summary_sheet.cell(row=row, column=5, value=_shape_text(record["arcrho_shape"]))
        summary_sheet.cell(row=row, column=6, value=_shape_text(record["resq_shape"]))
        if record["max_abs_diff"] is not None:
            summary_sheet.cell(row=row, column=7, value=record["max_abs_diff"]).number_format = NUMBER_FORMAT
        if record["max_rel_diff"] is not None:
            summary_sheet.cell(row=row, column=8, value=record["max_rel_diff"]).number_format = "0.00E+00"
        summary_sheet.cell(row=row, column=9, value=record["flagged_cells"] or None)
        summary_sheet.cell(row=row, column=10, value=record["note"])
        row += 1

    for rc_path, note in rc_errors:
        summary_sheet.cell(row=row, column=1, value=rc_path)
        summary_sheet.cell(row=row, column=2, value="(reserving class)")
        summary_sheet.cell(row=row, column=10, value=note)
        row += 1

    if row == 4:
        summary_sheet.cell(row=row, column=1, value="No dataset needs review.")

    _autosize(summary_sheet, min_width=12, max_width=80)

    coverage_sheet.cell(
        row=1,
        column=1,
        value=f"Every dataset compared, {len(records)} in total across {len(rc_paths)} reserving class(es).",
    ).font = styles["bold"]
    coverage_headers = ["RC Path", "Kind", "Dataset", "Source", "Cells", "Max Abs Diff", "Status"]
    for col, text in enumerate(coverage_headers, start=1):
        coverage_sheet.cell(row=3, column=col, value=text).font = styles["bold"]
    coverage_sheet.freeze_panes = "A4"
    row = 4
    for record in sorted(records, key=lambda r: (r["rc_path"], r["kind"], r["name"].casefold())):
        coverage_sheet.cell(row=row, column=1, value=record["rc_path"])
        coverage_sheet.cell(row=row, column=2, value=record["kind"])
        name_cell = coverage_sheet.cell(row=row, column=3, value=record["name"])
        anchor = anchors.get((record["rc_path"], record["kind"], record["name"]))
        if anchor:
            anchor_sheet_title, anchor_row = anchor
            name_cell.hyperlink = f"#'{anchor_sheet_title}'!A{anchor_row}"
            name_cell.font = styles["link"]
        coverage_sheet.cell(row=row, column=4, value=record["source_kind"])
        coverage_sheet.cell(row=row, column=5, value=record["row_count"] * record["column_count"] or None)
        if record["max_abs_diff"] is not None:
            coverage_sheet.cell(row=row, column=6, value=record["max_abs_diff"]).number_format = NUMBER_FORMAT
        coverage_sheet.cell(row=row, column=7, value="review" if record["needs_review"] else "match")
        row += 1
    _autosize(coverage_sheet, min_width=12, max_width=60)

    descriptor, temporary_name = tempfile.mkstemp(prefix=f".{path.stem}-", suffix=".xlsx", dir=path.parent)
    os.close(descriptor)
    temporary_path = Path(temporary_name)
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, path)
    finally:
        workbook.close()
        if temporary_path.exists():
            temporary_path.unlink()


def _parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--project", default=TARGET_PROJECT_NAME, help="ResQ project name to review.")
    parser.add_argument(
        "--source-kind",
        action="append",
        choices=[*SOURCE_KINDS, "all"],
        help="Dataset source kind to include; repeatable. Defaults to engine only.",
    )
    parser.add_argument(
        "--rc",
        action="append",
        help="Only review reserving classes whose path contains this text; repeatable.",
    )
    parser.add_argument("--no-open", action="store_true", help="Do not open the workbook when the run finishes.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = _parse_args(argv)
    selected = args.source_kind or list(DEFAULT_SOURCE_KINDS)
    source_kinds = SOURCE_KINDS if "all" in selected else tuple(dict.fromkeys(selected))

    rc_paths = RC_PATHS
    if args.rc:
        needles = [text.casefold() for text in args.rc]
        rc_paths = [path for path in RC_PATHS if any(needle in path.casefold() for needle in needles)]
        if not rc_paths:
            print("No reserving class matched --rc.")
            return 2

    records, rc_errors = run_comparison(
        project_name=args.project, rc_paths=rc_paths, source_kinds=source_kinds
    )
    output_path = _output_path(args.project)
    write_workbook(
        output_path,
        records,
        rc_errors,
        project_name=args.project,
        rc_paths=rc_paths,
        source_kinds=source_kinds,
    )
    review_records = [record for record in records if record["needs_review"]]
    needs_attention = bool(review_records) or bool(rc_errors)
    print(
        f"Compared {len(rc_paths)} reserving class(es), {len(records)} plain dataset(s) "
        f"[{', '.join(source_kinds)}]."
    )
    print(
        f"{len(review_records)} dataset(s) need review"
        + (f", {len(rc_errors)} reserving-class error(s)" if rc_errors else "")
        + "."
    )
    print(f"Excel report: {output_path}")
    if needs_attention and not args.no_open:
        try:
            os.startfile(output_path)  # noqa: S606 - opening the report just written, for the operator running this script
        except Exception as exc:
            print(f"Could not open the report automatically: {type(exc).__name__}: {exc}")
    return 0 if not needs_attention else 1


if __name__ == "__main__":
    raise SystemExit(main())
